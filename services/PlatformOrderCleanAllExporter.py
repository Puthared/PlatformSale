from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import database as db
from models.Platform import Platform
from models.PlatformOrder import PlatformOrder
from models.PlatformOrderItem import PlatformOrderItem
from models.ShopeeMaster import ShopeeMaster


MAX_EXPORT_ROWS = 50000
SUPPORTED_PLATFORMS = {"Shopee"}

CLEAN_ALL_HEADERS = [
    "หมายเลขคำสั่งซื้อออนไลน์",
    "วันที่สั่งซื้อ",
    "แพลตฟอร์ม",
    "SKU",
    "จำนวนเงินที่ควรได้รับ",
    "จังหวัด",
    "Order_Status_Clean",
    "Purchase_Hour",
    "Day_of_Week",
    "Month_Year",
    "Product_Name",
    "Basket Size",
    "Is_Unique_Order",
]


@dataclass(frozen=True)
class PlatformExportFilter:
    platform: str
    date_from: date
    date_to: date


@dataclass(frozen=True)
class CleanAllExportResult:
    file_path: Path
    file_name: str
    row_count: int


def export_clean_all_excel(filters: Iterable[PlatformExportFilter]) -> CleanAllExportResult:
    export_filters = list(filters)
    _validate_filters(export_filters)

    rows = _build_clean_all_rows(export_filters)
    if len(rows) > MAX_EXPORT_ROWS:
        raise ValueError(
            f"Export result has {len(rows)} rows. Please reduce date range to stay under {MAX_EXPORT_ROWS} rows."
        )

    file_name = f"Clean_ALL_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.xlsx"
    export_dir = Path.cwd() / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / file_name

    _write_workbook(file_path, rows)
    return CleanAllExportResult(file_path=file_path, file_name=file_name, row_count=len(rows))


def _validate_filters(filters: list[PlatformExportFilter]) -> None:
    if not filters:
        raise ValueError("Please select at least one platform.")

    for item in filters:
        platform = item.platform.strip()
        if platform not in SUPPORTED_PLATFORMS:
            raise ValueError(f"Platform '{platform}' is not supported yet.")
        if item.date_from > item.date_to:
            raise ValueError(f"date_from cannot be greater than date_to for {platform}.")


def _build_clean_all_rows(filters: list[PlatformExportFilter]) -> list[list]:
    clean_rows: list[list] = []

    for item in filters:
        if item.platform.strip() == "Shopee":
            clean_rows.extend(_build_shopee_rows(item))

    return clean_rows


def _build_shopee_rows(item: PlatformExportFilter) -> list[list]:
    start_at = datetime.combine(item.date_from, time.min)
    end_at = datetime.combine(item.date_to, time.max)

    query_rows = (
        db.session
        .query(PlatformOrder, PlatformOrderItem, Platform.PlatformName, ShopeeMaster.ShippingProvince)
        .join(Platform, PlatformOrder.PlatformId == Platform.PlatformId)
        .join(PlatformOrderItem, PlatformOrder.PlatformOrderId == PlatformOrderItem.PlatformOrderId)
        .outerjoin(ShopeeMaster, PlatformOrder.RawSourceId == ShopeeMaster.ShopeeMasterId)
        .filter(Platform.PlatformName == "Shopee")
        .filter(Platform.isDeleted == False)
        .filter(PlatformOrder.isDeleted == False)
        .filter(PlatformOrderItem.isDeleted == False)
        .filter(PlatformOrder.OrderCreatedAt >= start_at)
        .filter(PlatformOrder.OrderCreatedAt <= end_at)
        .order_by(PlatformOrder.OrderCreatedAt, PlatformOrder.PlatformOrderId, PlatformOrderItem.PlatformOrderItemId)
        .all()
    )

    clean_rows: list[list] = []
    seen_orders: set[int] = set()

    for order, order_item, platform_name, shipping_province in query_rows:
        is_first_order_row = order.PlatformOrderId not in seen_orders
        seen_orders.add(order.PlatformOrderId)

        total_amount = _decimal_to_float(order.TotalAmount)
        order_created_at = order.OrderCreatedAt

        clean_rows.append(
            [
                order.PlatformOrderNo if is_first_order_row else "",
                order_created_at.date() if is_first_order_row and order_created_at else "",
                platform_name if is_first_order_row else "",
                order_item.SellerSku or order_item.PlatformSku or "",
                total_amount if is_first_order_row else "",
                shipping_province if is_first_order_row else "",
                order.OrderStatus if is_first_order_row else "",
                f"{order_created_at.hour:02d}" if is_first_order_row and order_created_at else "",
                order_created_at.strftime("%A") if is_first_order_row and order_created_at else "",
                date(order_created_at.year, order_created_at.month, 1) if is_first_order_row and order_created_at else "",
                order_item.ProductName or "",
                _basket_size(total_amount) if is_first_order_row else "",
                1 if is_first_order_row else 0,
            ]
        )

    return clean_rows


def _decimal_to_float(value: Decimal | None) -> float:
    if value is None:
        return 0
    return float(value)


def _basket_size(total_amount: float) -> str:
    if total_amount <= 500:
        return "0 - 500"
    if total_amount <= 1000:
        return "501 - 1,000"
    if total_amount <= 2000:
        return "1,001 - 2,000"
    if total_amount <= 3000:
        return "2,001 - 3,000"
    return "> 3,000"


def _write_workbook(file_path: Path, rows: list[list]) -> None:
    workbook = Workbook(write_only=False)
    worksheet = workbook.active
    worksheet.title = "Clean_ALL"

    worksheet.append(CLEAN_ALL_HEADERS)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, header in enumerate(CLEAN_ALL_HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = _column_width(header, rows, column_index)

    for cell in worksheet["E"][1:]:
        cell.number_format = '#,##0.00'

    for cell in worksheet["B"][1:]:
        cell.number_format = "yyyy-mm-dd"

    for cell in worksheet["J"][1:]:
        cell.number_format = "yyyy-mm-dd"

    workbook.save(file_path)


def _column_width(header: str, rows: list[list], column_index: int) -> int:
    max_length = len(str(header))
    for row in rows[:200]:
        value = row[column_index - 1]
        max_length = max(max_length, len(str(value)) if value is not None else 0)
    return min(max(max_length + 2, 12), 42)
