from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FILE_PATH = Path(__file__).resolve().parent / "ImuraAnalysis.xlsx"
SHEET_NAME = "Note"


def add_section(ws, row, number, title, objective, displays):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row, 1, f"{number}. {title}")
    cell.font = Font(size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 25

    ws.cell(row + 1, 1, "เป้าหมาย")
    ws.cell(row + 1, 2, objective)
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=4)

    ws.cell(row + 2, 1, "สิ่งที่ต้องแสดง")
    for index, display in enumerate(displays, start=row + 2):
        if index > row + 2:
            ws.cell(index, 1, "")
        ws.cell(index, 2, f"• {display}")
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=4)

    end_row = row + 1 + len(displays)
    for current_row in range(row + 1, end_row + 1):
        ws.cell(current_row, 1).font = Font(bold=True, color="1F1F1F")
        ws.cell(current_row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        for column in range(1, 5):
            ws.cell(current_row, column).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        ws.row_dimensions[current_row].height = 32
    return end_row + 2


def main():
    workbook = load_workbook(FILE_PATH)
    if SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[SHEET_NAME])
    worksheet = workbook.create_sheet(SHEET_NAME, 0)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Imura Marketing Analysis Plan"
    worksheet["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 38

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = (
        "แผนวิเคราะห์ประสิทธิภาพการขายจาก Shopee, TikTok และ Lazada "
        "โดยใช้ข้อมูลจาก Data Imura.xlsx / Clean_All"
    )
    worksheet["A2"].font = Font(italic=True, color="44546A")
    worksheet["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 32

    worksheet.merge_cells("A4:D4")
    worksheet["A4"] = "กติกาการนับข้อมูล"
    worksheet["A4"].font = Font(size=14, bold=True, color="FFFFFF")
    worksheet["A4"].fill = PatternFill("solid", fgColor="548235")

    rules = [
        ("จำนวน Orders", "นับเฉพาะแถว Is_Unique_Order = 1 หรือแถวที่มีหมายเลขคำสั่งซื้อ"),
        ("Revenue", "รวม Revenue ซึ่งถูกบันทึกเพียงครั้งเดียวต่อ Order"),
        ("All Orders", "ใช้เพื่อวัด Demand รวมทุกสถานะ"),
        ("Completed Orders", "ใช้เพื่อวัดยอดขายสำเร็จจริง"),
        ("สถานะอื่น", "แสดง Cancelled, Shipping และ Returned แยกต่างหาก"),
    ]
    current_row = 5
    for label, detail in rules:
        worksheet.cell(current_row, 1, label)
        worksheet.cell(current_row, 2, detail)
        worksheet.merge_cells(
            start_row=current_row, start_column=2, end_row=current_row, end_column=4
        )
        worksheet.cell(current_row, 1).font = Font(bold=True)
        worksheet.cell(current_row, 1).fill = PatternFill("solid", fgColor="E2F0D9")
        worksheet.cell(current_row, 2).alignment = Alignment(wrap_text=True)
        current_row += 1

    current_row += 2
    sections = [
        (
            "Executive Summary",
            "ตอบทันทีว่าแพลตฟอร์มใดขายดีที่สุดในแต่ละมิติ",
            [
                "KPI แยก Shopee, TikTok และ Lazada: Total Orders, Completed Orders, Revenue และ Completed Revenue",
                "Average Order Value (AOV), Revenue Share และ Order Share",
                "Cancellation Rate และ Return Rate",
                "จัดอันดับแพลตฟอร์มที่ Orders, Revenue และ AOV สูงที่สุด รวมถึงแพลตฟอร์มที่มีอัตรายกเลิกต่ำที่สุด",
            ],
        ),
        (
            "Daily Performance",
            "เปรียบเทียบยอดขายและจำนวน Orders ระดับวันต่อวัน",
            [
                "กราฟ Revenue รายวันแยกแพลตฟอร์ม",
                "กราฟ Orders และ AOV รายวันแยกแพลตฟอร์ม",
                "อัตราเติบโตเทียบวันก่อนหน้า และเทียบวันเดียวกันของสัปดาห์ก่อน",
                "ตาราง Daily Comparison พร้อมระบุแพลตฟอร์มที่ชนะในแต่ละวัน",
            ],
        ),
        (
            "Campaign Day Analysis",
            "วัดว่าวันแคมเปญช่วยเพิ่ม Orders และ Revenue จริงหรือไม่",
            [
                "จัดประเภท Double Day, วันที่ 15, Payday, สิ้นเดือน และวันแคมเปญที่กำหนดเอง",
                "เปรียบเทียบค่าเฉลี่ย Orders, Revenue, AOV และ Cancellation Rate ระหว่างวันแคมเปญกับวันปกติ",
                "แสดง Uplift เป็นจำนวนและเปอร์เซ็นต์",
                "เปรียบเทียบกับวันปกติที่เป็นวันในสัปดาห์เดียวกัน เพื่อลดผลกระทบจากวันหยุด",
            ],
        ),
        (
            "Hourly Analysis",
            "ค้นหาช่วงเวลาที่เหมาะสำหรับจัดโปรโมชัน ไลฟ์ และยิงโฆษณา",
            [
                "Heatmap เดือน × ชั่วโมง, แพลตฟอร์ม × ชั่วโมง และวันในสัปดาห์ × ชั่วโมง",
                "วิเคราะห์ Orders, Revenue และ AOV รายชั่วโมง",
                "ระบุชั่วโมงขายดีที่สุดของแต่ละแพลตฟอร์มในแต่ละเดือน",
                "สรุปช่วงเวลาที่ควรเพิ่มงบโฆษณาหรือจัดกิจกรรมการขาย",
            ],
        ),
        (
            "Status And Sales Quality",
            "ประเมินคุณภาพยอดขาย ไม่ตัดสินจาก Revenue เพียงอย่างเดียว",
            [
                "Completed Rate, Cancellation Rate และ Return Rate แยกแพลตฟอร์ม",
                "จำนวน Orders ที่ยังอยู่ระหว่าง Shipping หรือ Processing",
                "Revenue ที่ยังไม่ Completed และรายได้ที่สูญเสียจาก Cancelled Orders",
                "เปรียบเทียบแพลตฟอร์มที่ขายดีแต่มีความเสี่ยงจากการยกเลิกหรือคืนสินค้าสูง",
            ],
        ),
    ]

    for number, section in enumerate(sections, start=1):
        current_row = add_section(
            worksheet,
            current_row,
            number,
            section[0],
            section[1],
            section[2],
        )

    thin = Side(style="thin", color="B7C9D6")
    for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(bottom=thin)

    widths = {1: 24, 2: 38, 3: 30, 4: 30}
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width

    workbook.save(FILE_PATH)
    print(f"Created {SHEET_NAME} in {FILE_PATH}")


if __name__ == "__main__":
    main()
