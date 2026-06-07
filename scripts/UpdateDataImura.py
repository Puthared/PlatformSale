from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_TARGET = SCRIPT_DIR / "Data Imura.xlsx"
DEFAULT_SHOPEE = SCRIPT_DIR / "ImuraMasterShopee.xlsx"
DEFAULT_LAZADA = SCRIPT_DIR / "ImuraMasterLazada.xlsx"
DEFAULT_TIKTOK = SCRIPT_DIR / "ImuraMasterTiktok.xlsx"

SOURCE_SHEETS = [
    ("Clean_Shopee", DEFAULT_SHOPEE, "Shopee"),
    ("Clean_Tiktok", DEFAULT_TIKTOK, "TikTok"),
    ("Clean_Lazada", DEFAULT_LAZADA, "Lazada"),
]

CLEAN_ALL_HEADERS = [
    "หมายเลขคำสั่งซื้อออนไลน์",
    "วันที่สั่งซื้อ",
    "แพลตฟอร์ม",
    "SKU",
    "Revenue",
    "จังหวัด",
    "Order_Status_Clean",
    "Purchase_Hour",
    "Day_of_Week",
    "Month_Year",
    "Product_Name",
    "Basket Size",
    "Is_Unique_Order",
    "ตัวเลือกสินค้า",
]


def to_number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def basket_size(value: Any) -> str:
    amount = to_number(value)
    if amount <= 800:
        return "<= 800"
    if amount <= 1500:
        return "801 - 1,500"
    if amount <= 2000:
        return "1,501 - 2,000"
    if amount <= 3000:
        return "2,001 - 3,000"
    return "> 3,000"


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ Sheet {sheet_name} ใน {path}")
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    rows = [list(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return headers, rows


def records_from(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    return [dict(zip(headers, row)) for row in rows]


def style_sheet(worksheet, widths: dict[int, float] | None = None) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for column_index in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = (
            widths.get(column_index, 18) if widths else 18
        )


def replace_sheet_data(workbook, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    if sheet_name in workbook.sheetnames:
        index = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook[sheet_name])
        worksheet = workbook.create_sheet(sheet_name, index)
    else:
        worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    style_sheet(worksheet, {4: 18, 11: 55})
    if worksheet.max_column >= 5:
        for cell in worksheet["E"][1:]:
            cell.number_format = '#,##0.00'
    if worksheet.max_column >= 10:
        for cell in worksheet["J"][1:]:
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd"


def clean_all_row(record: dict[str, Any], platform: str) -> list[Any]:
    revenue = to_number(record.get("จำนวนเงินที่ควรได้รับ"))
    clean_status = record.get("Order_Status_Clean")
    if platform == "Lazada" and clean_status == "confirmed":
        clean_status = "Completed"

    product_name = record.get("Product_Name")
    if product_name in (None, ""):
        product_name = record.get("Product_NAME")

    source_basket = record.get("Basket Size")
    output_basket = source_basket if source_basket not in (None, "") else basket_size(revenue)

    unique_order = record.get("Is_Unique_Order")
    if unique_order in (None, ""):
        unique_order = 1

    option = (
        record.get("ตัวเลือกสินค้า")
        or record.get("ตัวเลือก")
        or record.get("ชื่อตัวเลือก")
        or None
    )

    return [
        record.get("หมายเลขคำสั่งซื้อออนไลน์"),
        record.get("วันที่สั่งซื้อ"),
        platform,
        record.get("SKU"),
        revenue,
        record.get("จังหวัด"),
        clean_status,
        record.get("Purchase_Hour"),
        record.get("Day_of_Week"),
        record.get("Month_Year"),
        product_name,
        output_basket,
        unique_order,
        option,
    ]


def update_data_imura(
    target: Path,
    source_paths: dict[str, Path],
    create_backup: bool = True,
) -> tuple[Path | None, dict[str, dict[str, float]]]:
    source_data: dict[str, tuple[list[str], list[list[Any]]]] = {}
    summary: dict[str, dict[str, float]] = {}

    for sheet_name, _, platform in SOURCE_SHEETS:
        headers, rows = read_sheet(source_paths[sheet_name], sheet_name)
        source_data[sheet_name] = (headers, rows)
        records = records_from(headers, rows)
        summary[platform] = {
            "rows": len(rows),
            "orders": sum(
                record.get("หมายเลขคำสั่งซื้อออนไลน์") not in (None, "")
                for record in records
            ),
            "revenue": sum(to_number(record.get("จำนวนเงินที่ควรได้รับ")) for record in records),
        }

    if not target.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์เป้าหมาย {target}")

    backup_path = None
    if create_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = target.with_name(f"{target.stem}.backup_{stamp}{target.suffix}")
        shutil.copy2(target, backup_path)

    workbook = load_workbook(target)
    for sheet_name, _, _ in SOURCE_SHEETS:
        headers, rows = source_data[sheet_name]
        replace_sheet_data(workbook, sheet_name, headers, rows)

    all_rows: list[list[Any]] = []
    for sheet_name, _, platform in SOURCE_SHEETS:
        headers, rows = source_data[sheet_name]
        all_rows.extend(clean_all_row(record, platform) for record in records_from(headers, rows))

    replace_sheet_data(workbook, "Clean_All", CLEAN_ALL_HEADERS, all_rows)
    workbook.save(target)

    summary["All"] = {
        "rows": len(all_rows),
        "orders": sum(values["orders"] for values in summary.values()),
        "revenue": sum(values["revenue"] for values in summary.values()),
    }
    return backup_path, summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh Data Imura.xlsx from all platform master files.")
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    parser.add_argument("--shopee", type=Path, default=DEFAULT_SHOPEE)
    parser.add_argument("--lazada", type=Path, default=DEFAULT_LAZADA)
    parser.add_argument("--tiktok", type=Path, default=DEFAULT_TIKTOK)
    parser.add_argument("--no-backup", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sources = {
        "Clean_Shopee": args.shopee,
        "Clean_Lazada": args.lazada,
        "Clean_Tiktok": args.tiktok,
    }
    backup, summary = update_data_imura(args.target, sources, not args.no_backup)
    print("Data Imura update completed")
    if backup:
        print(f"Backup: {backup}")
    for platform, values in summary.items():
        print(
            f"{platform}: rows={int(values['rows'])}, "
            f"orders={int(values['orders'])}, revenue={values['revenue']:.2f}"
        )


if __name__ == "__main__":
    main()
