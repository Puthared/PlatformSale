from __future__ import annotations

import argparse
from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_RAW_FILE = SCRIPT_DIR / "LD_1-7_June.xlsx"
DEFAULT_MASTER_FILE = SCRIPT_DIR / "ImuraMasterLazada.xlsx"

RAW_SHEET = "Raw_Lazada"
MASTER_ORDERS_SHEET = "Master_Orders"
MASTER_ITEMS_SHEET = "Master_Items"
CLEAN_SHEET = "Clean_Lazada"
IMPORT_LOG_SHEET = "Import_Log"

ORDER_ID = "orderNumber"

CLEAN_HEADERS = [
    "หมายเลขคำสั่งซื้อออนไลน์",
    "วันที่สั่งซื้อ",
    "แพลตฟอร์ม",
    "SKU",
    "จำนวนเงินที่ควรได้รับ",
    "จังหวัด",
    "Order_Status_Clean",
    "Purchase_Hour",
    "Day_of_Week",
    "Month_Year",
    "Product_NAME",
]

MASTER_ORDER_HEADERS = [
    "orderNumber",
    "orderType",
    "deliveryType",
    "createTime",
    "updateTime",
    "deliveredDate",
    "shippingCity",
    "shippingRegion",
    "billingCity",
    "shippingPostCode",
    "shippingCountry",
    "payMethod",
    "status",
    "shippingProvider",
    "trackingCode",
    "Item_Line_Count",
    "Total_Paid_Price",
]

MASTER_ITEM_HEADERS = [
    "orderNumber",
    "Item_Line",
    "orderItemId",
    "lazadaId",
    "sellerSku",
    "lazadaSku",
    "paidPrice",
    "unitPrice",
    "sellerDiscountTotal",
    "platformDiscountTotal",
    "shippingFee",
    "walletCredit",
    "itemName",
    "variation",
    "status",
    "bundleId",
    "bundleDiscount",
    "refundAmount",
]

DATE_FORMATS = (
    "%d %b %Y %H:%M",
    "%d %B %Y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
)


def to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_decimal(value: Any) -> Decimal:
    if value in (None, "", "-"):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def decimal_value(value: Any) -> float:
    return float(to_decimal(value))


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = to_text(value)
    if not text or text == "-":
        return None
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    return None


def normalize_status(value: Any) -> str:
    status = to_text(value)
    mapping = {
        "confirmed": "Completed",
        "delivered": "Completed",
        "shipped": "Completed",
        "ready_to_ship": "Processing",
        "pending": "Processing",
        "canceled": "Cancelled",
        "cancelled": "Cancelled",
        "returned": "Returned",
        "failed": "Failed",
    }
    return mapping.get(status, status)


def province_from(order: dict[str, Any]) -> str:
    return (
        to_text(order.get("shippingRegion"))
        or to_text(order.get("shippingCity"))
        or to_text(order.get("billingCity"))
    )


def read_raw_orders(raw_file: Path) -> OrderedDict[str, dict[str, Any]]:
    workbook = load_workbook(raw_file, read_only=True, data_only=True)
    if RAW_SHEET not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ Sheet {RAW_SHEET} ใน {raw_file}")

    worksheet = workbook[RAW_SHEET]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    required = {
        "orderNumber",
        "orderItemId",
        "sellerSku",
        "createTime",
        "paidPrice",
        "itemName",
        "status",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Raw_Lazada ขาด Field: {', '.join(missing)}")

    orders: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        order_id = to_text(record.get(ORDER_ID))
        if not order_id:
            continue

        if order_id not in orders:
            order = {header: record.get(header) for header in MASTER_ORDER_HEADERS[:-2]}
            order["Item_Line_Count"] = 0
            order["Total_Paid_Price"] = Decimal("0")
            orders[order_id] = {"Order": order, "Items": []}

        item = {header: record.get(header) for header in MASTER_ITEM_HEADERS[2:]}
        item[ORDER_ID] = order_id
        item["Item_Line"] = len(orders[order_id]["Items"]) + 1
        orders[order_id]["Items"].append(item)
        orders[order_id]["Order"]["Item_Line_Count"] += 1
        orders[order_id]["Order"]["Total_Paid_Price"] += to_decimal(record.get("paidPrice"))

    workbook.close()
    return orders


def rows_to_dicts(worksheet) -> list[dict[str, Any]]:
    if worksheet.max_row < 2:
        return []
    headers = [cell.value for cell in worksheet[1]]
    return [
        dict(zip(headers, values))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]


def load_existing_orders(master_file: Path) -> OrderedDict[str, dict[str, Any]]:
    orders: OrderedDict[str, dict[str, Any]] = OrderedDict()
    if not master_file.exists():
        return orders

    workbook = load_workbook(master_file, read_only=True, data_only=True)
    if MASTER_ORDERS_SHEET in workbook.sheetnames:
        for order in rows_to_dicts(workbook[MASTER_ORDERS_SHEET]):
            order_id = to_text(order.get(ORDER_ID))
            if order_id:
                orders[order_id] = {"Order": order, "Items": []}

    if MASTER_ITEMS_SHEET in workbook.sheetnames:
        for item in rows_to_dicts(workbook[MASTER_ITEMS_SHEET]):
            order_id = to_text(item.get(ORDER_ID))
            if order_id in orders:
                orders[order_id]["Items"].append(item)

    workbook.close()
    return orders


def open_master_workbook(master_file: Path) -> Workbook:
    workbook = load_workbook(master_file) if master_file.exists() else Workbook()
    for default_name in ("Sheet", "Sheet1"):
        if workbook.sheetnames == [default_name] and workbook[default_name].max_row == 1:
            workbook.remove(workbook[default_name])
    return workbook


def replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        index = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook[sheet_name])
        return workbook.create_sheet(sheet_name, index)
    return workbook.create_sheet(sheet_name)


def style_sheet(worksheet, widths: dict[int, float] | None = None) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for column_index in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = (
            widths.get(column_index, 18) if widths else 18
        )


def write_rows(worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet.append(headers)
    for record in rows:
        worksheet.append(
            [
                decimal_value(record.get(header))
                if isinstance(record.get(header), Decimal)
                else record.get(header)
                for header in headers
            ]
        )


def clean_rows(orders: OrderedDict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for order_id, group in orders.items():
        order = group["Order"]
        first_item = group["Items"][0] if group["Items"] else {}
        created = parse_datetime(order.get("createTime"))
        result.append(
            {
                "หมายเลขคำสั่งซื้อออนไลน์": order_id,
                "วันที่สั่งซื้อ": created or order.get("createTime"),
                "แพลตฟอร์ม": "Lazada",
                "SKU": first_item.get("sellerSku", ""),
                "จำนวนเงินที่ควรได้รับ": decimal_value(order.get("Total_Paid_Price")),
                "จังหวัด": province_from(order),
                "Order_Status_Clean": normalize_status(order.get("status")),
                "Purchase_Hour": created.hour if created else "",
                "Day_of_Week": created.strftime("%A") if created else "",
                "Month_Year": datetime(created.year, created.month, 1) if created else "",
                "Product_NAME": first_item.get("itemName", ""),
            }
        )
    return result


def append_import_log(workbook: Workbook, summary: dict[str, Any]) -> None:
    headers = [
        "Imported_At",
        "Source_File",
        "Raw_Item_Rows",
        "Imported_Orders",
        "Inserted_Orders",
        "Updated_Orders",
        "Master_Orders",
        "Master_Items",
        "Total_Revenue",
    ]
    if IMPORT_LOG_SHEET not in workbook.sheetnames:
        worksheet = workbook.create_sheet(IMPORT_LOG_SHEET)
        worksheet.append(headers)
    else:
        worksheet = workbook[IMPORT_LOG_SHEET]
    worksheet.append([summary.get(header) for header in headers])
    style_sheet(worksheet)


def build_master(raw_file: Path, master_file: Path) -> dict[str, Any]:
    imported = read_raw_orders(raw_file)
    existing = load_existing_orders(master_file)
    inserted = sum(order_id not in existing for order_id in imported)
    updated = len(imported) - inserted

    for order_id, group in imported.items():
        existing[order_id] = group

    workbook = open_master_workbook(master_file)
    order_rows = [group["Order"] for group in existing.values()]
    item_rows = [item for group in existing.values() for item in group["Items"]]
    generated_clean_rows = clean_rows(existing)

    worksheet = replace_sheet(workbook, MASTER_ORDERS_SHEET)
    write_rows(worksheet, MASTER_ORDER_HEADERS, order_rows)
    style_sheet(worksheet)

    worksheet = replace_sheet(workbook, MASTER_ITEMS_SHEET)
    write_rows(worksheet, MASTER_ITEM_HEADERS, item_rows)
    style_sheet(worksheet, {13: 55, 14: 25})

    worksheet = replace_sheet(workbook, CLEAN_SHEET)
    write_rows(worksheet, CLEAN_HEADERS, generated_clean_rows)
    style_sheet(worksheet, {4: 18, 11: 55})
    for cell in worksheet["B"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"
    for cell in worksheet["E"][1:]:
        cell.number_format = '#,##0.00'
    for cell in worksheet["J"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"

    total_revenue = sum(
        (to_decimal(group["Order"].get("Total_Paid_Price")) for group in existing.values()),
        Decimal("0"),
    )
    summary = {
        "Imported_At": datetime.now(),
        "Source_File": str(raw_file),
        "Raw_Item_Rows": sum(len(group["Items"]) for group in imported.values()),
        "Imported_Orders": len(imported),
        "Inserted_Orders": inserted,
        "Updated_Orders": updated,
        "Master_Orders": len(existing),
        "Master_Items": len(item_rows),
        "Total_Revenue": float(total_revenue),
    }
    append_import_log(workbook, summary)
    master_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(master_file)
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Raw_Lazada into a reusable Lazada master workbook.")
    parser.add_argument("--raw-file", type=Path, default=DEFAULT_RAW_FILE)
    parser.add_argument("--master-file", type=Path, default=DEFAULT_MASTER_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.raw_file.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ Raw: {args.raw_file}")
    summary = build_master(args.raw_file, args.master_file)
    print("Lazada import completed")
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
