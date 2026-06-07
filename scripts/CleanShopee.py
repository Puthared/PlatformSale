from __future__ import annotations

import argparse
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_RAW_FILE = SCRIPT_DIR / "SP_1-7 June.xlsx"
DEFAULT_MASTER_FILE = SCRIPT_DIR / "ImuraMasterShopee.xlsx"

RAW_SHEET = "Raw_Shopee"
MASTER_ORDERS_SHEET = "Master_Orders"
MASTER_ITEMS_SHEET = "Master_Items"
CLEAN_SHEET = "Clean_Shopee"
IMPORT_LOG_SHEET = "Import_Log"

ORDER_ID = "หมายเลขคำสั่งซื้อ"

CLEAN_HEADERS = [
    "หมายเลขคำสั่งซื้อออนไลน์",
    "วันที่สั่งซื้อ",
    "แพลตฟอร์ม",
    "SKU",
    "จำนวนเงินที่ควรได้รับ",
    "จังหวัด",
    "Order_Status_Clean",
    "Purchase_Hour",
    "Day_of_Week",
    "Month_Year",
    "Product_Name",
    "Basket Size",
    "Is_Unique_Order",
]

MASTER_ORDER_HEADERS = [
    ORDER_ID,
    "สถานะการสั่งซื้อ",
    "เหตุผลในการยกเลิกคำสั่งซื้อ",
    "สถานะการคืนเงินหรือคืนสินค้า",
    "ชื่อผู้ใช้ (ผู้ซื้อ)",
    "วันที่ทำการสั่งซื้อ",
    "เวลาการชำระสินค้า",
    "ช่องทางการชำระเงิน",
    "ตัวเลือกการจัดส่ง",
    "วิธีการจัดส่ง",
    "*หมายเลขติดตามพัสดุ",
    "จำนวนเงินทั้งหมด",
    "จังหวัด",
    "เขต/อำเภอ",
    "รหัสไปรษณีย์",
    "เวลาที่ทำการสั่งซื้อสำเร็จ",
    "Item_Line_Count",
    "Total_Quantity",
]

MASTER_ITEM_HEADERS = [
    ORDER_ID,
    "Item_Line",
    "เลขอ้างอิง Parent SKU",
    "เลขอ้างอิง SKU (SKU Reference No.)",
    "ชื่อสินค้า",
    "ชื่อตัวเลือก",
    "ราคาตั้งต้น",
    "ราคาขาย",
    "จำนวน",
    "จำนวนที่ส่งคืน",
    "ราคาขายสุทธิ",
    "ส่วนลดจาก Shopee",
    "โค้ดส่วนลดชำระโดยผู้ขาย",
    "โค้ดส่วนลดชำระโดย Shopee (เช่น โค้ดจากโปรแกรม ร้านโค้ดคุ้ม, โค้ดส่วนลด Shopee, โค้ดส่วนลด Shopee Mall)",
    "ค่าคอมมิชชั่น",
    "Transaction Fee",
    "ราคาสินค้าที่ชำระโดยผู้ซื้อ (THB)",
]

DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
)


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = to_text(value)
    if not text or text == "-":
        return None
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    return None


def normalize_status(raw_status: Any) -> str:
    status = to_text(raw_status)
    if status == "สำเร็จแล้ว":
        return "Completed"
    if status == "ยกเลิกแล้ว":
        return "Cancelled"
    if status == "จัดส่งสำเร็จแล้ว":
        return "Completed"
    if status in {"การจัดส่ง", "ที่ต้องจัดส่ง"}:
        return "Shipping"
    if status.startswith("ผู้ซื้อได้รับสินค้าแล้ว โปรดทราบว่า"):
        return "DeliveredPendingReturn"
    return status


def basket_size(amount: Any) -> str:
    value = to_number(amount)
    if value <= 500:
        return "0 - 500"
    if value <= 1000:
        return "501 - 1,000"
    if value <= 2000:
        return "1,001 - 2,000"
    if value <= 3000:
        return "2,001 - 3,000"
    return "> 3,000"


def selected_sku(item: dict[str, Any]) -> str:
    return to_text(item.get("เลขอ้างอิง Parent SKU")) or to_text(
        item.get("เลขอ้างอิง SKU (SKU Reference No.)")
    )


def read_raw_orders(raw_file: Path) -> OrderedDict[str, dict[str, Any]]:
    workbook = load_workbook(raw_file, read_only=True, data_only=True)
    if RAW_SHEET not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ Sheet {RAW_SHEET} ใน {raw_file}")

    worksheet = workbook[RAW_SHEET]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    required = {
        ORDER_ID,
        "สถานะการสั่งซื้อ",
        "วันที่ทำการสั่งซื้อ",
        "เลขอ้างอิง Parent SKU",
        "เลขอ้างอิง SKU (SKU Reference No.)",
        "ชื่อสินค้า",
        "จำนวน",
        "จำนวนเงินทั้งหมด",
        "จังหวัด",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Raw_Shopee ขาด Field: {', '.join(missing)}")

    orders: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        order_id = to_text(record.get(ORDER_ID))
        if not order_id:
            continue

        if order_id not in orders:
            order_values = {header: record.get(header) for header in MASTER_ORDER_HEADERS[:-2]}
            order_values["Item_Line_Count"] = 0
            order_values["Total_Quantity"] = 0.0
            orders[order_id] = {"Order": order_values, "Items": []}

        item = {header: record.get(header) for header in MASTER_ITEM_HEADERS[2:]}
        item[ORDER_ID] = order_id
        item["Item_Line"] = len(orders[order_id]["Items"]) + 1
        orders[order_id]["Items"].append(item)
        orders[order_id]["Order"]["Item_Line_Count"] += 1
        orders[order_id]["Order"]["Total_Quantity"] += to_number(record.get("จำนวน"))

    workbook.close()
    return orders


def rows_to_dicts(worksheet) -> list[dict[str, Any]]:
    if worksheet.max_row < 2:
        return []
    headers = [cell.value for cell in worksheet[1]]
    return [
        dict(zip(headers, values))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]


def load_existing_orders(master_file: Path) -> OrderedDict[str, dict[str, Any]]:
    orders: OrderedDict[str, dict[str, Any]] = OrderedDict()
    if not master_file.exists():
        return orders

    workbook = load_workbook(master_file, read_only=True, data_only=True)
    if MASTER_ORDERS_SHEET in workbook.sheetnames:
        for order in rows_to_dicts(workbook[MASTER_ORDERS_SHEET]):
            order_id = to_text(order.get(ORDER_ID))
            if order_id:
                orders[order_id] = {"Order": order, "Items": []}

    if MASTER_ITEMS_SHEET in workbook.sheetnames:
        for item in rows_to_dicts(workbook[MASTER_ITEMS_SHEET]):
            order_id = to_text(item.get(ORDER_ID))
            if order_id in orders:
                orders[order_id]["Items"].append(item)

    workbook.close()
    return orders


def open_master_workbook(master_file: Path) -> Workbook:
    if master_file.exists():
        workbook = load_workbook(master_file)
    else:
        workbook = Workbook()
    if workbook.sheetnames == ["Sheet"] and workbook["Sheet"].max_row == 1:
        workbook.remove(workbook["Sheet"])
    if workbook.sheetnames == ["Sheet1"] and workbook["Sheet1"].max_row == 1:
        workbook.remove(workbook["Sheet1"])
    return workbook


def replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        index = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook[sheet_name])
        return workbook.create_sheet(sheet_name, index)
    return workbook.create_sheet(sheet_name)


def style_sheet(worksheet, widths: dict[int, float] | None = None) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for column_index in range(1, worksheet.max_column + 1):
        width = widths.get(column_index, 18) if widths else 18
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def write_rows(worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet.append(headers)
    for record in rows:
        worksheet.append([record.get(header) for header in headers])


def clean_rows(orders: OrderedDict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for order_id, group in orders.items():
        order = group["Order"]
        created = parse_datetime(order.get("วันที่ทำการสั่งซื้อ"))
        items = group["Items"] or [{}]
        for item_index, item in enumerate(items):
            first = item_index == 0
            result.append(
                {
                    "หมายเลขคำสั่งซื้อออนไลน์": order_id if first else "",
                    "วันที่สั่งซื้อ": created if first else "",
                    "แพลตฟอร์ม": "Shopee" if first else "",
                    "SKU": selected_sku(item),
                    "จำนวนเงินที่ควรได้รับ": to_number(order.get("จำนวนเงินทั้งหมด")) if first else "",
                    "จังหวัด": order.get("จังหวัด") if first else "",
                    "Order_Status_Clean": normalize_status(order.get("สถานะการสั่งซื้อ")) if first else "",
                    "Purchase_Hour": created.strftime("%H") if first and created else "",
                    "Day_of_Week": created.strftime("%A") if first and created else "",
                    "Month_Year": datetime(created.year, created.month, 1) if first and created else "",
                    "Product_Name": item.get("ชื่อสินค้า", ""),
                    "Basket Size": basket_size(order.get("จำนวนเงินทั้งหมด")) if first else "",
                    "Is_Unique_Order": 1 if first else 0,
                }
            )
    return result


def append_import_log(workbook: Workbook, summary: dict[str, Any]) -> None:
    headers = [
        "Imported_At",
        "Source_File",
        "Raw_Item_Rows",
        "Imported_Orders",
        "Inserted_Orders",
        "Updated_Orders",
        "Master_Orders",
        "Master_Items",
        "Total_Revenue",
    ]
    if IMPORT_LOG_SHEET not in workbook.sheetnames:
        worksheet = workbook.create_sheet(IMPORT_LOG_SHEET)
        worksheet.append(headers)
    else:
        worksheet = workbook[IMPORT_LOG_SHEET]
    worksheet.append([summary.get(header) for header in headers])
    style_sheet(worksheet)


def build_master(raw_file: Path, master_file: Path) -> dict[str, Any]:
    imported = read_raw_orders(raw_file)
    existing = load_existing_orders(master_file)
    inserted = sum(1 for order_id in imported if order_id not in existing)
    updated = len(imported) - inserted

    for order_id, group in imported.items():
        existing[order_id] = group

    workbook = open_master_workbook(master_file)

    order_rows = [group["Order"] for group in existing.values()]
    item_rows = [item for group in existing.values() for item in group["Items"]]
    generated_clean_rows = clean_rows(existing)

    orders_sheet = replace_sheet(workbook, MASTER_ORDERS_SHEET)
    write_rows(orders_sheet, MASTER_ORDER_HEADERS, order_rows)
    style_sheet(orders_sheet)

    items_sheet = replace_sheet(workbook, MASTER_ITEMS_SHEET)
    write_rows(items_sheet, MASTER_ITEM_HEADERS, item_rows)
    style_sheet(items_sheet, {5: 55, 6: 25})

    clean_sheet = replace_sheet(workbook, CLEAN_SHEET)
    write_rows(clean_sheet, CLEAN_HEADERS, generated_clean_rows)
    style_sheet(clean_sheet, {4: 18, 11: 55})
    for cell in clean_sheet["B"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"
    for cell in clean_sheet["E"][1:]:
        cell.number_format = '#,##0.00'
    for cell in clean_sheet["J"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"

    summary = {
        "Imported_At": datetime.now(),
        "Source_File": str(raw_file),
        "Raw_Item_Rows": sum(len(group["Items"]) for group in imported.values()),
        "Imported_Orders": len(imported),
        "Inserted_Orders": inserted,
        "Updated_Orders": updated,
        "Master_Orders": len(existing),
        "Master_Items": len(item_rows),
        "Total_Revenue": sum(
            to_number(group["Order"].get("จำนวนเงินทั้งหมด")) for group in existing.values()
        ),
    }
    append_import_log(workbook, summary)

    master_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(master_file)
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Raw_Shopee into a reusable Shopee master workbook.")
    parser.add_argument("--raw-file", type=Path, default=DEFAULT_RAW_FILE)
    parser.add_argument("--master-file", type=Path, default=DEFAULT_MASTER_FILE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.raw_file.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ Raw: {args.raw_file}")
    summary = build_master(args.raw_file, args.master_file)
    print("Shopee import completed")
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
