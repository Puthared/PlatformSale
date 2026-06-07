from __future__ import annotations

import shutil
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE_FILE = SCRIPT_DIR / "Data Imura.xlsx"
ANALYSIS_FILE = SCRIPT_DIR / "ImuraAnalysis.xlsx"
SOURCE_SHEET = "Clean_All"
DAILY_SHEET = "Daily Performance"
DATA_SHEET = "Analysis_Daily"

PLATFORMS = ["Shopee", "TikTok", "Lazada"]
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

NAVY = "17365D"
BLUE = "1F4E78"
GREEN = "548235"
ORANGE = "C65911"
RED = "C00000"
WHITE = "FFFFFF"
GRAY = "F2F2F2"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_ORANGE = "FCE4D6"


def number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def read_records() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    worksheet = workbook[SOURCE_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    records = [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    workbook.close()
    return records


def pct_change(current: float, previous: float) -> float | None:
    return (current - previous) / previous if previous else None


def aggregate_daily(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    daily = defaultdict(
        lambda: defaultdict(
            lambda: {
                "orders": 0,
                "revenue": 0.0,
                "completed_orders": 0,
                "completed_revenue": 0.0,
                "cancelled_orders": 0,
            }
        )
    )
    for record in records:
        order_date = parse_date(record.get("วันที่สั่งซื้อ"))
        platform = record.get("แพลตฟอร์ม")
        if not order_date or platform not in PLATFORMS:
            continue
        values = daily[order_date][platform]
        revenue = number(record.get("Revenue"))
        values["revenue"] += revenue
        if record.get("หมายเลขคำสั่งซื้อออนไลน์") not in (None, ""):
            values["orders"] += 1
            if record.get("Order_Status_Clean") == "Completed":
                values["completed_orders"] += 1
            if record.get("Order_Status_Clean") == "Cancelled":
                values["cancelled_orders"] += 1
        if record.get("Order_Status_Clean") == "Completed":
            values["completed_revenue"] += revenue

    result = []
    sorted_dates = sorted(daily)
    by_date = {}
    for order_date in sorted_dates:
        platform_values = daily[order_date]
        total_orders = sum(platform_values[p]["orders"] for p in PLATFORMS)
        total_revenue = sum(platform_values[p]["revenue"] for p in PLATFORMS)
        total_completed_orders = sum(platform_values[p]["completed_orders"] for p in PLATFORMS)
        total_completed_revenue = sum(platform_values[p]["completed_revenue"] for p in PLATFORMS)
        total_cancelled = sum(platform_values[p]["cancelled_orders"] for p in PLATFORMS)
        row = {
            "date": order_date,
            "day": order_date.strftime("%A"),
            "orders": total_orders,
            "revenue": total_revenue,
            "completed_orders": total_completed_orders,
            "completed_revenue": total_completed_revenue,
            "aov": total_revenue / total_orders if total_orders else 0,
            "completed_rate": total_completed_orders / total_orders if total_orders else 0,
            "cancel_rate": total_cancelled / total_orders if total_orders else 0,
            "revenue_winner": max(PLATFORMS, key=lambda p: platform_values[p]["revenue"]),
            "order_winner": max(PLATFORMS, key=lambda p: platform_values[p]["orders"]),
        }
        for platform in PLATFORMS:
            row[f"{platform}_orders"] = platform_values[platform]["orders"]
            row[f"{platform}_revenue"] = platform_values[platform]["revenue"]
        by_date[order_date] = row
        result.append(row)

    for row in result:
        previous = by_date.get(row["date"].fromordinal(row["date"].toordinal() - 1))
        previous_week = by_date.get(row["date"].fromordinal(row["date"].toordinal() - 7))
        row["revenue_dod"] = pct_change(row["revenue"], previous["revenue"]) if previous else None
        row["orders_dod"] = pct_change(row["orders"], previous["orders"]) if previous else None
        row["revenue_wow"] = pct_change(row["revenue"], previous_week["revenue"]) if previous_week else None
        row["orders_wow"] = pct_change(row["orders"], previous_week["orders"]) if previous_week else None
    return result


def replace_sheet(workbook, name: str, index: int | None = None):
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    return workbook.create_sheet(name, index)


def style_header(cell, fill=NAVY):
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_card(ws, cell_range: str, title: str, value: str, fill: str):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = f"{title}\n{value}"
    cell.font = Font(size=13, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def section_title(ws, row: int, title: str, fill: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row, 1, title)
    cell.font = Font(size=15, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)


def build_data_sheet(workbook, daily_rows):
    ws = replace_sheet(workbook, DATA_SHEET)
    headers = [
        "วันที่", "วัน", "Order ทั้งหมด", "Revenue ทั้งหมด", "Completed Order",
        "Completed Revenue", "มูลค่าเฉลี่ยต่อ Order", "อัตราสำเร็จ", "อัตรายกเลิก",
        "Revenue เทียบวันก่อน", "Order เทียบวันก่อน", "Revenue เทียบ 7 วันก่อน",
        "Order เทียบ 7 วันก่อน", "แพลตฟอร์ม Revenue สูงสุด", "แพลตฟอร์ม Order สูงสุด",
        "Shopee Order", "TikTok Order", "Lazada Order",
        "Shopee Revenue", "TikTok Revenue", "Lazada Revenue",
    ]
    ws.append(headers)
    for row in daily_rows:
        ws.append([
            row["date"], row["day"], row["orders"], row["revenue"],
            row["completed_orders"], row["completed_revenue"], row["aov"],
            row["completed_rate"], row["cancel_rate"], row["revenue_dod"],
            row["orders_dod"], row["revenue_wow"], row["orders_wow"],
            row["revenue_winner"], row["order_winner"],
            row["Shopee_orders"], row["TikTok_orders"], row["Lazada_orders"],
            row["Shopee_revenue"], row["TikTok_revenue"], row["Lazada_revenue"],
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        style_header(cell, BLUE)
    for cell in ws["A"][1:]:
        cell.number_format = "dd-mmm-yyyy"
    for column in (4, 6, 7, 19, 20, 21):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00'
    for column in (8, 9, 10, 11, 12, 13):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = '0.0%'
    ws.sheet_state = "hidden"


def build_dashboard(workbook, daily_rows):
    ws = replace_sheet(workbook, DAILY_SHEET, 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    for column in range(1, 17):
        ws.column_dimensions[get_column_letter(column)].width = 13

    ws.merge_cells("A1:P1")
    ws["A1"] = "IMURA Daily Performance"
    ws["A1"].font = Font(size=22, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:P2")
    ws["A2"] = (
        f"ช่วงข้อมูล {daily_rows[0]['date']:%d %b %Y} - {daily_rows[-1]['date']:%d %b %Y}"
        f"  |  อัปเดตแดชบอร์ด {datetime.now():%d %b %Y %H:%M}"
    )
    ws["A2"].font = Font(italic=True, color="44546A")
    ws["A2"].alignment = Alignment(horizontal="center")

    total_revenue = sum(row["revenue"] for row in daily_rows)
    total_orders = sum(row["orders"] for row in daily_rows)
    best_revenue = max(daily_rows, key=lambda row: row["revenue"])
    best_orders = max(daily_rows, key=lambda row: row["orders"])
    cards = [
        ("A4:D6", "Revenue เฉลี่ยต่อวัน", f"{total_revenue / len(daily_rows):,.2f} บาท", BLUE),
        ("E4:H6", "Order เฉลี่ยต่อวัน", f"{total_orders / len(daily_rows):,.1f}", GREEN),
        ("I4:L6", "วันที่ Revenue สูงสุด", f"{best_revenue['date']:%d %b %Y}\n{best_revenue['revenue']:,.2f} บาท", ORANGE),
        ("M4:P6", "วันที่ Order สูงสุด", f"{best_orders['date']:%d %b %Y}\n{best_orders['orders']:,.0f} Order", GREEN),
    ]
    for card in cards:
        add_card(ws, *card)

    section_title(ws, 8, "Daily Winner และวิเคราะห์วันในสัปดาห์", GREEN)
    revenue_winners = Counter(row["revenue_winner"] for row in daily_rows)
    order_winners = Counter(row["order_winner"] for row in daily_rows)
    weekday_values = defaultdict(list)
    for row in daily_rows:
        weekday_values[row["day"]].append(row)

    source_col = 18
    ws.cell(2, source_col, "แพลตฟอร์ม")
    ws.cell(2, source_col + 1, "วันที่ Revenue สูงสุด")
    ws.cell(2, source_col + 2, "วันที่ Order สูงสุด")
    for row_index, platform in enumerate(PLATFORMS, 3):
        ws.cell(row_index, source_col, platform)
        ws.cell(row_index, source_col + 1, revenue_winners[platform])
        ws.cell(row_index, source_col + 2, order_winners[platform])

    ws.cell(8, source_col, "วัน")
    ws.cell(8, source_col + 1, "Revenue เฉลี่ย")
    ws.cell(8, source_col + 2, "Order เฉลี่ย")
    for row_index, day_name in enumerate(DAY_NAMES, 9):
        values = weekday_values[day_name]
        ws.cell(row_index, source_col, day_name)
        ws.cell(row_index, source_col + 1, sum(row["revenue"] for row in values) / len(values))
        ws.cell(row_index, source_col + 2, sum(row["orders"] for row in values) / len(values))

    ws.merge_cells("A9:H9")
    ws["A9"] = (
        f"จำนวนวันที่แต่ละแพลตฟอร์มทำ Revenue หรือ Order สูงที่สุด "
        f"จากวันที่มีข้อมูลทั้งหมด {len(daily_rows):,} วัน"
    )
    ws["A9"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws["A9"].font = Font(italic=True, color="44546A")

    ws.merge_cells("I9:P9")
    ws["I9"] = "Revenue เฉลี่ยต่อวันของแต่ละวันในสัปดาห์ ใช้ดูว่าวันใดสร้างรายได้ได้ดีที่สุดโดยเฉลี่ย"
    ws["I9"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws["I9"].font = Font(italic=True, color="44546A")
    ws.row_dimensions[9].height = 32

    winner_chart = BarChart()
    winner_chart.type = "col"
    winner_chart.title = "จำนวนวันที่แต่ละแพลตฟอร์มทำยอดสูงสุด"
    winner_chart.y_axis.title = "จำนวนวันที่ชนะ"
    winner_chart.x_axis.title = "แพลตฟอร์ม"
    winner_chart.height = 9
    winner_chart.width = 14
    winner_chart.add_data(Reference(ws, min_col=source_col + 1, max_col=source_col + 2, min_row=2, max_row=5), titles_from_data=True)
    winner_chart.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=5))
    winner_chart.dataLabels = DataLabelList()
    winner_chart.dataLabels.showVal = True
    winner_chart.legend.position = "b"
    winner_chart.visible_cells_only = False
    ws.add_chart(winner_chart, "A11")

    weekday_chart = BarChart()
    weekday_chart.type = "bar"
    weekday_chart.title = "Revenue เฉลี่ยต่อวัน แยกตามวันในสัปดาห์"
    weekday_chart.x_axis.title = "Revenue เฉลี่ย (บาท)"
    weekday_chart.y_axis.title = "วันในสัปดาห์"
    weekday_chart.height = 9
    weekday_chart.width = 14
    weekday_chart.add_data(Reference(ws, min_col=source_col + 1, min_row=8, max_row=15), titles_from_data=True)
    weekday_chart.set_categories(Reference(ws, min_col=source_col, min_row=9, max_row=15))
    weekday_chart.dataLabels = DataLabelList()
    weekday_chart.dataLabels.showVal = True
    weekday_chart.legend = None
    weekday_chart.visible_cells_only = False
    ws.add_chart(weekday_chart, "I11")

    section_title(ws, 32, "เปรียบเทียบผลการขายรายวัน", ORANGE)
    headers = [
        "วันที่", "วัน", "Order", "Revenue", "Completed Revenue", "มูลค่าเฉลี่ยต่อ Order",
        "อัตราสำเร็จ", "Revenue เทียบวันก่อน", "Order เทียบวันก่อน",
        "Revenue เทียบ 7 วันก่อน", "Order เทียบ 7 วันก่อน",
        "แพลตฟอร์ม Revenue สูงสุด", "แพลตฟอร์ม Order สูงสุด",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(33, col, header)
        style_header(ws.cell(33, col), NAVY)
    for row_index, row in enumerate(reversed(daily_rows), 34):
        values = [
            row["date"], row["day"], row["orders"], row["revenue"], row["completed_revenue"],
            row["aov"], row["completed_rate"], row["revenue_dod"], row["orders_dod"],
            row["revenue_wow"], row["orders_wow"], row["revenue_winner"], row["order_winner"],
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row_index, col, value)
        ws.cell(row_index, 1).number_format = "dd-mmm-yyyy"
        for col in (4, 5, 6):
            ws.cell(row_index, col).number_format = '#,##0.00'
        for col in (7, 8, 9, 10, 11):
            ws.cell(row_index, col).number_format = '0.0%'

    ws.auto_filter.ref = f"A33:M{33 + len(daily_rows)}"
    thin = Side(style="thin", color="B7C9D6")
    for row in ws.iter_rows(min_row=33, max_row=33 + len(daily_rows), min_col=1, max_col=13):
        for cell in row:
            cell.border = Border(bottom=thin)
    for col in (1, 2, 12, 13):
        ws.column_dimensions[get_column_letter(col)].width = 20

    section_row = 35 + len(daily_rows)
    section_title(ws, section_row, "สรุป", ORANGE)
    platform_revenue_avg = {}
    for platform in PLATFORMS:
        platform_revenue_avg[platform] = sum(row[f"{platform}_revenue"] for row in daily_rows) / len(daily_rows)
    best_platform = max(PLATFORMS, key=lambda p: platform_revenue_avg[p])
    best_weekday = max(DAY_NAMES, key=lambda d: sum(r["revenue"] for r in weekday_values[d]) / len(weekday_values[d]))
    insights = [
        f"{best_platform} มี Revenue เฉลี่ยต่อวันสูงที่สุด {platform_revenue_avg[best_platform]:,.2f} บาท",
        f"{best_revenue['date']:%d %b %Y} เป็นวันที่มี Revenue รวมสูงที่สุด {best_revenue['revenue']:,.2f} บาท",
        f"{best_orders['date']:%d %b %Y} เป็นวันที่มี Order รวมสูงที่สุด {best_orders['orders']:,.0f} Order",
        f"{best_weekday} เป็นวันในสัปดาห์ที่มี Revenue เฉลี่ยสูงที่สุด",
        f"{max(PLATFORMS, key=lambda p: revenue_winners[p])} เป็นแพลตฟอร์มที่ชนะด้าน Revenue บ่อยที่สุด {max(revenue_winners.values())} วัน",
    ]
    for row_index, insight in enumerate(insights, section_row + 1):
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=16)
        ws.cell(row_index, 1, f"• {insight}")
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_ORANGE if row_index % 2 else GRAY)
        ws.cell(row_index, 1).alignment = Alignment(wrap_text=True)

    for column in range(source_col, source_col + 4):
        ws.column_dimensions[get_column_letter(column)].hidden = True


def main():
    records = read_records()
    daily_rows = aggregate_daily(records)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ANALYSIS_FILE.with_name(f"{ANALYSIS_FILE.stem}.backup_{stamp}{ANALYSIS_FILE.suffix}")
    shutil.copy2(ANALYSIS_FILE, backup)
    workbook = load_workbook(ANALYSIS_FILE)
    build_data_sheet(workbook, daily_rows)
    build_dashboard(workbook, daily_rows)
    workbook.save(ANALYSIS_FILE)
    print(f"Updated: {ANALYSIS_FILE}")
    print(f"Backup: {backup}")
    print(f"Days: {len(daily_rows)}")


if __name__ == "__main__":
    main()
