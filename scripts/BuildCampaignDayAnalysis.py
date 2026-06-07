from __future__ import annotations

import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE_FILE = SCRIPT_DIR / "Data Imura.xlsx"
ANALYSIS_FILE = SCRIPT_DIR / "ImuraAnalysis.xlsx"
SOURCE_SHEET = "Clean_All"
OUTPUT_SHEET = "Campaign Day Analysis"
DATA_SHEET = "Analysis_Campaign"

PLATFORMS = ["Shopee", "TikTok", "Lazada"]
CAMPAIGN_TYPES = ["Double Day", "Mid-Month", "Payday", "Month Start", "Normal Day"]
CAMPAIGN_ONLY = CAMPAIGN_TYPES[:-1]

NAVY = "17365D"
BLUE = "1F4E78"
GREEN = "548235"
ORANGE = "C65911"
RED = "C00000"
WHITE = "FFFFFF"
GRAY = "F2F2F2"
LIGHT_ORANGE = "FCE4D6"


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def campaign_type(day: date) -> str:
    if day.day == day.month:
        return "Double Day"
    if day.day == 15:
        return "Mid-Month"
    if day.day >= 25:
        return "Payday"
    if day.day <= 3:
        return "Month Start"
    return "Normal Day"


def read_records() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    worksheet = workbook[SOURCE_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    records = [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    workbook.close()
    return records


def new_metrics() -> dict[str, float]:
    return {
        "orders": 0,
        "revenue": 0.0,
        "completed_orders": 0,
        "completed_revenue": 0.0,
        "cancelled_orders": 0,
        "returned_orders": 0,
    }


def aggregate_daily(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    daily = defaultdict(lambda: defaultdict(new_metrics))
    for record in records:
        order_date = parse_date(record.get("วันที่สั่งซื้อ"))
        platform = record.get("แพลตฟอร์ม")
        if not order_date or platform not in PLATFORMS:
            continue
        metric = daily[order_date][platform]
        revenue = number(record.get("Revenue"))
        status = record.get("Order_Status_Clean")
        metric["revenue"] += revenue
        if record.get("หมายเลขคำสั่งซื้อออนไลน์") not in (None, ""):
            metric["orders"] += 1
            metric["completed_orders"] += status == "Completed"
            metric["cancelled_orders"] += status == "Cancelled"
            metric["returned_orders"] += status in {"Returned", "Package Returned", "Lost by 3PL"}
        if status == "Completed":
            metric["completed_revenue"] += revenue

    result = []
    for order_date in sorted(daily):
        platforms = daily[order_date]
        total = new_metrics()
        for platform in PLATFORMS:
            for key in total:
                total[key] += platforms[platform][key]
        total.update(
            {
                "date": order_date,
                "day": order_date.strftime("%A"),
                "campaign_type": campaign_type(order_date),
                "platform_winner": max(PLATFORMS, key=lambda p: platforms[p]["revenue"]),
                "platforms": platforms,
            }
        )
        result.append(total)
    return result


def summarize_days(days: list[dict[str, Any]]) -> dict[str, float]:
    count = len(days)
    orders = sum(day["orders"] for day in days)
    revenue = sum(day["revenue"] for day in days)
    completed_orders = sum(day["completed_orders"] for day in days)
    cancelled_orders = sum(day["cancelled_orders"] for day in days)
    returned_orders = sum(day["returned_orders"] for day in days)
    return {
        "days": count,
        "orders": orders,
        "revenue": revenue,
        "completed_revenue": sum(day["completed_revenue"] for day in days),
        "avg_orders": orders / count if count else 0,
        "avg_revenue": revenue / count if count else 0,
        "avg_completed_revenue": sum(day["completed_revenue"] for day in days) / count if count else 0,
        "aov": revenue / orders if orders else 0,
        "completed_rate": completed_orders / orders if orders else 0,
        "cancel_rate": cancelled_orders / orders if orders else 0,
        "return_rate": returned_orders / orders if orders else 0,
    }


def pct_change(value: float, base: float) -> float | None:
    return (value - base) / base if base else None


def replace_sheet(workbook, name: str, index: int | None = None):
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    return workbook.create_sheet(name, index)


def style_header(cell, fill=NAVY):
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def section_title(ws, row: int, title: str, fill: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row, 1, title)
    cell.font = Font(size=15, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)


def add_card(ws, cell_range: str, title: str, value: str, fill: str):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = f"{title}\n{value}"
    cell.font = Font(size=13, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_sheet(workbook, daily_rows, summaries, platform_summaries):
    ws = replace_sheet(workbook, DATA_SHEET)
    headers = [
        "วันที่", "วัน", "ประเภทวัน", "Order", "Revenue", "Completed Revenue",
        "AOV", "อัตราสำเร็จ", "อัตรายกเลิก", "อัตราคืนสินค้า", "แพลตฟอร์ม Revenue สูงสุด",
    ]
    ws.append(headers)
    for day in daily_rows:
        summary = summarize_days([day])
        ws.append([
            day["date"], day["day"], day["campaign_type"], day["orders"], day["revenue"],
            day["completed_revenue"], summary["aov"], summary["completed_rate"],
            summary["cancel_rate"], summary["return_rate"], day["platform_winner"],
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        style_header(cell, BLUE)
    for cell in ws["A"][1:]:
        cell.number_format = "dd-mmm-yyyy"
    ws.sheet_state = "hidden"


def build_dashboard(workbook, daily_rows, summaries, platform_summaries):
    ws = replace_sheet(workbook, OUTPUT_SHEET, 2)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    for column in range(1, 17):
        ws.column_dimensions[get_column_letter(column)].width = 13

    normal = summaries["Normal Day"]
    campaign_days = [day for day in daily_rows if day["campaign_type"] != "Normal Day"]
    campaign = summarize_days(campaign_days)
    best_campaign_day = max(campaign_days, key=lambda day: day["revenue"])

    ws.merge_cells("A1:P1")
    ws["A1"] = "IMURA Campaign Day Analysis"
    ws["A1"].font = Font(size=22, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38
    ws.merge_cells("A2:P2")
    ws["A2"] = (
        f"ช่วงข้อมูล {daily_rows[0]['date']:%d %b %Y} - {daily_rows[-1]['date']:%d %b %Y}"
        f"  |  อัปเดตแดชบอร์ด {datetime.now():%d %b %Y %H:%M}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="44546A")

    revenue_uplift = pct_change(campaign["avg_revenue"], normal["avg_revenue"]) or 0
    order_uplift = pct_change(campaign["avg_orders"], normal["avg_orders"]) or 0
    aov_change = pct_change(campaign["aov"], normal["aov"]) or 0
    cards = [
        ("A4:D6", "จำนวนวัน Campaign", f"{campaign['days']:,.0f} วัน", BLUE),
        ("E4:H6", "Revenue เฉลี่ยวัน Campaign", f"{campaign['avg_revenue']:,.2f} บาท", GREEN),
        ("I4:L6", "Revenue Uplift เทียบ Normal Day", f"{revenue_uplift:+.1%}", ORANGE),
        ("M4:P6", "Order Uplift เทียบ Normal Day", f"{order_uplift:+.1%}", GREEN),
        ("A8:D10", "Revenue วัน Normal เฉลี่ย", f"{normal['avg_revenue']:,.2f} บาท", BLUE),
        ("E8:H10", "AOV Change", f"{aov_change:+.1%}", GREEN),
        ("I8:L10", "Campaign Revenue รวม", f"{campaign['revenue']:,.2f} บาท", ORANGE),
        ("M8:P10", "Campaign ที่ Revenue สูงสุด", f"{best_campaign_day['date']:%d %b %Y}\n{best_campaign_day['revenue']:,.2f} บาท", GREEN),
    ]
    for card in cards:
        add_card(ws, *card)

    section_title(ws, 12, "เปรียบเทียบประเภทวัน Campaign กับ Normal Day", BLUE)
    headers = [
        "ประเภทวัน", "จำนวนวัน", "Order เฉลี่ย", "Revenue เฉลี่ย", "Completed Revenue เฉลี่ย",
        "มูลค่าเฉลี่ยต่อ Order", "อัตราสำเร็จ", "อัตรายกเลิก", "อัตราคืนสินค้า",
        "Revenue Uplift", "Order Uplift", "AOV Change",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(13, col, header)
        style_header(ws.cell(13, col))
    for row_index, name in enumerate(CAMPAIGN_TYPES, 14):
        value = summaries[name]
        values = [
            name, value["days"], value["avg_orders"], value["avg_revenue"],
            value["avg_completed_revenue"], value["aov"], value["completed_rate"],
            value["cancel_rate"], value["return_rate"],
            pct_change(value["avg_revenue"], normal["avg_revenue"]) if name != "Normal Day" else 0,
            pct_change(value["avg_orders"], normal["avg_orders"]) if name != "Normal Day" else 0,
            pct_change(value["aov"], normal["aov"]) if name != "Normal Day" else 0,
        ]
        for col, item in enumerate(values, 1):
            ws.cell(row_index, col, item)
        for col in (3, 4, 5, 6):
            ws.cell(row_index, col).number_format = '#,##0.00'
        for col in (7, 8, 9, 10, 11, 12):
            ws.cell(row_index, col).number_format = '0.0%'

    source_col = 18
    ws.cell(2, source_col, "ประเภทวัน")
    ws.cell(2, source_col + 1, "Revenue Uplift")
    ws.cell(2, source_col + 2, "Order Uplift")
    for row_index, name in enumerate(CAMPAIGN_ONLY, 3):
        ws.cell(row_index, source_col, name)
        ws.cell(row_index, source_col + 1, pct_change(summaries[name]["avg_revenue"], normal["avg_revenue"]))
        ws.cell(row_index, source_col + 2, pct_change(summaries[name]["avg_orders"], normal["avg_orders"]))

    ws.merge_cells("A20:H20")
    ws["A20"] = "Revenue Uplift แสดงการเปลี่ยนแปลงของ Revenue เฉลี่ยต่อวัน เมื่อเทียบกับ Normal Day"
    ws["A20"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("I20:P20")
    ws["I20"] = "Order Uplift แสดงการเปลี่ยนแปลงของจำนวน Order เฉลี่ยต่อวัน เมื่อเทียบกับ Normal Day"
    ws["I20"].alignment = Alignment(horizontal="center", wrap_text=True)

    revenue_chart = BarChart()
    revenue_chart.type = "col"
    revenue_chart.title = "Revenue Uplift เทียบ Normal Day"
    revenue_chart.y_axis.title = "Uplift"
    revenue_chart.x_axis.title = "ประเภท Campaign"
    revenue_chart.height = 8
    revenue_chart.width = 14
    revenue_chart.add_data(Reference(ws, min_col=source_col + 1, min_row=2, max_row=6), titles_from_data=True)
    revenue_chart.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=6))
    revenue_chart.dataLabels = DataLabelList()
    revenue_chart.dataLabels.showVal = True
    revenue_chart.legend = None
    revenue_chart.visible_cells_only = False
    ws.add_chart(revenue_chart, "A21")

    order_chart = BarChart()
    order_chart.type = "col"
    order_chart.title = "Order Uplift เทียบ Normal Day"
    order_chart.y_axis.title = "Uplift"
    order_chart.x_axis.title = "ประเภท Campaign"
    order_chart.height = 8
    order_chart.width = 14
    order_chart.add_data(Reference(ws, min_col=source_col + 2, min_row=2, max_row=6), titles_from_data=True)
    order_chart.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=6))
    order_chart.dataLabels = DataLabelList()
    order_chart.dataLabels.showVal = True
    order_chart.legend = None
    order_chart.visible_cells_only = False
    ws.add_chart(order_chart, "I21")

    section_title(ws, 38, "ผล Campaign แยกตามแพลตฟอร์ม", GREEN)
    platform_headers = [
        "ประเภทวัน", "แพลตฟอร์ม", "จำนวนวัน", "Order เฉลี่ย", "Revenue เฉลี่ย",
        "มูลค่าเฉลี่ยต่อ Order", "อัตราสำเร็จ", "อัตรายกเลิก", "Revenue Uplift", "Order Uplift",
    ]
    for col, header in enumerate(platform_headers, 1):
        ws.cell(39, col, header)
        style_header(ws.cell(39, col))
    row_index = 40
    for name in CAMPAIGN_ONLY:
        for platform in PLATFORMS:
            value = platform_summaries[name][platform]
            base = platform_summaries["Normal Day"][platform]
            values = [
                name, platform, value["days"], value["avg_orders"], value["avg_revenue"],
                value["aov"], value["completed_rate"], value["cancel_rate"],
                pct_change(value["avg_revenue"], base["avg_revenue"]),
                pct_change(value["avg_orders"], base["avg_orders"]),
            ]
            for col, item in enumerate(values, 1):
                ws.cell(row_index, col, item)
            for col in (4, 5, 6):
                ws.cell(row_index, col).number_format = '#,##0.00'
            for col in (7, 8, 9, 10):
                ws.cell(row_index, col).number_format = '0.0%'
            row_index += 1

    section_title(ws, 54, "ผล Double Day แต่ละเดือน", ORANGE)
    double_days = [day for day in daily_rows if day["campaign_type"] == "Double Day"]
    double_headers = [
        "Campaign", "วัน", "Order", "Revenue", "Completed Revenue", "มูลค่าเฉลี่ยต่อ Order",
        "อัตราสำเร็จ", "อัตรายกเลิก", "แพลตฟอร์ม Revenue สูงสุด",
    ]
    for col, header in enumerate(double_headers, 1):
        ws.cell(55, col, header)
        style_header(ws.cell(55, col))
    for index, day in enumerate(double_days, 56):
        single = summarize_days([day])
        values = [
            day["date"].strftime("%d.%m.%Y"), day["day"], day["orders"], day["revenue"],
            day["completed_revenue"], single["aov"], single["completed_rate"],
            single["cancel_rate"], day["platform_winner"],
        ]
        for col, item in enumerate(values, 1):
            ws.cell(index, col, item)
        for col in (4, 5, 6):
            ws.cell(index, col).number_format = '#,##0.00'
        for col in (7, 8):
            ws.cell(index, col).number_format = '0.0%'

    calendar_start = 58 + len(double_days)
    section_title(ws, calendar_start, "Campaign Calendar", BLUE)
    calendar_headers = [
        "วันที่", "วัน", "ประเภทวัน", "Order", "Revenue", "Completed Revenue",
        "มูลค่าเฉลี่ยต่อ Order", "อัตราสำเร็จ", "อัตรายกเลิก", "แพลตฟอร์ม Revenue สูงสุด",
    ]
    for col, header in enumerate(calendar_headers, 1):
        ws.cell(calendar_start + 1, col, header)
        style_header(ws.cell(calendar_start + 1, col))
    for index, day in enumerate(reversed(daily_rows), calendar_start + 2):
        single = summarize_days([day])
        values = [
            day["date"], day["day"], day["campaign_type"], day["orders"], day["revenue"],
            day["completed_revenue"], single["aov"], single["completed_rate"],
            single["cancel_rate"], day["platform_winner"],
        ]
        for col, item in enumerate(values, 1):
            ws.cell(index, col, item)
        ws.cell(index, 1).number_format = "dd-mmm-yyyy"
        for col in (5, 6, 7):
            ws.cell(index, col).number_format = '#,##0.00'
        for col in (8, 9):
            ws.cell(index, col).number_format = '0.0%'
    calendar_end = calendar_start + 1 + len(daily_rows)
    ws.auto_filter.ref = f"A{calendar_start + 1}:J{calendar_end}"
    ws.conditional_formatting.add(
        f"E{calendar_start + 2}:E{calendar_end}",
        ColorScaleRule(start_type="min", start_color="F4CCCC", mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color="D9EAD3"),
    )

    summary_row = calendar_end + 2
    section_title(ws, summary_row, "สรุป", ORANGE)
    best_type = max(CAMPAIGN_ONLY, key=lambda name: summaries[name]["avg_revenue"])
    best_platform_type = max(
        ((name, platform) for name in CAMPAIGN_ONLY for platform in PLATFORMS),
        key=lambda item: pct_change(
            platform_summaries[item[0]][item[1]]["avg_revenue"],
            platform_summaries["Normal Day"][item[1]]["avg_revenue"],
        ) or -999,
    )
    insights = [
        f"{best_type} มี Revenue เฉลี่ยต่อวันสูงที่สุด {summaries[best_type]['avg_revenue']:,.2f} บาท",
        f"วัน Campaign มี Revenue เฉลี่ยสูงกว่า Normal Day {revenue_uplift:+.1%}",
        f"วัน Campaign มีจำนวน Order เฉลี่ยสูงกว่า Normal Day {order_uplift:+.1%}",
        f"{best_platform_type[1]} ได้รับ Revenue Uplift จาก {best_platform_type[0]} สูงที่สุด",
        f"{best_campaign_day['date']:%d %b %Y} เป็นวัน Campaign ที่มี Revenue สูงที่สุด {best_campaign_day['revenue']:,.2f} บาท",
    ]
    for index, insight in enumerate(insights, summary_row + 1):
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=16)
        ws.cell(index, 1, f"• {insight}")
        ws.cell(index, 1).fill = PatternFill("solid", fgColor=LIGHT_ORANGE if index % 2 else GRAY)
        ws.cell(index, 1).alignment = Alignment(wrap_text=True)

    for column in range(source_col, source_col + 4):
        ws.column_dimensions[get_column_letter(column)].hidden = True
    thin = Side(style="thin", color="B7C9D6")
    for row in ws.iter_rows(min_row=13, max_row=calendar_end, min_col=1, max_col=12):
        for cell in row:
            if cell.value is not None:
                cell.border = Border(bottom=thin)


def main():
    records = read_records()
    daily_rows = aggregate_daily(records)
    grouped = {name: [day for day in daily_rows if day["campaign_type"] == name] for name in CAMPAIGN_TYPES}
    summaries = {name: summarize_days(days) for name, days in grouped.items()}

    platform_summaries = defaultdict(dict)
    for name, days in grouped.items():
        for platform in PLATFORMS:
            platform_days = []
            for day in days:
                values = dict(day["platforms"][platform])
                values["date"] = day["date"]
                platform_days.append(values)
            platform_summaries[name][platform] = summarize_days(platform_days)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ANALYSIS_FILE.with_name(f"{ANALYSIS_FILE.stem}.backup_{stamp}{ANALYSIS_FILE.suffix}")
    shutil.copy2(ANALYSIS_FILE, backup)
    workbook = load_workbook(ANALYSIS_FILE)
    write_data_sheet(workbook, daily_rows, summaries, platform_summaries)
    build_dashboard(workbook, daily_rows, summaries, platform_summaries)
    workbook.save(ANALYSIS_FILE)
    print(f"Updated: {ANALYSIS_FILE}")
    print(f"Backup: {backup}")
    for name in CAMPAIGN_TYPES:
        print(f"{name}: days={summaries[name]['days']}, avg_revenue={summaries[name]['avg_revenue']:.2f}")


if __name__ == "__main__":
    main()
