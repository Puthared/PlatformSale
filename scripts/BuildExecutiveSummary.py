from __future__ import annotations

import shutil
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE_FILE = SCRIPT_DIR / "Data Imura.xlsx"
ANALYSIS_FILE = SCRIPT_DIR / "ImuraAnalysis.xlsx"
SOURCE_SHEET = "Clean_All"
DATA_SHEET = "Analysis_Data"
SUMMARY_SHEET = "Executive Summary"

PLATFORMS = ["Shopee", "TikTok", "Lazada"]
AT_RISK_STATUSES = {"Shipping", "Processing", "DeliveredPendingReturn", "ยังไม่ชำระ"}
RETURN_STATUSES = {"Returned", "Package Returned", "Lost by 3PL"}

NAVY = "17365D"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GREEN = "548235"
LIGHT_GREEN = "E2F0D9"
ORANGE = "C65911"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "F2F2F2"
WHITE = "FFFFFF"


def number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            pass
    return None


def read_clean_all() -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    worksheet = workbook[SOURCE_SHEET]
    headers = [cell.value for cell in worksheet[1]]
    records = [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    workbook.close()
    return headers, records


def calculate_metrics(records: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    metrics: dict[str, dict[str, Any]] = {}
    for platform in PLATFORMS:
        platform_records = [row for row in records if row.get("แพลตฟอร์ม") == platform]
        orders = [row for row in platform_records if row.get("หมายเลขคำสั่งซื้อออนไลน์") not in (None, "")]
        status_counts = Counter(row.get("Order_Status_Clean") for row in orders)
        revenue = sum(number(row.get("Revenue")) for row in platform_records)
        completed_revenue = sum(
            number(row.get("Revenue"))
            for row in platform_records
            if row.get("Order_Status_Clean") == "Completed"
        )
        at_risk_revenue = sum(
            number(row.get("Revenue"))
            for row in platform_records
            if row.get("Order_Status_Clean") in AT_RISK_STATUSES
        )
        cancelled_revenue = sum(
            number(row.get("Revenue"))
            for row in platform_records
            if row.get("Order_Status_Clean") == "Cancelled"
        )
        total_orders = len(orders)
        completed_orders = status_counts["Completed"]
        cancelled_orders = status_counts["Cancelled"]
        returned_orders = sum(status_counts[status] for status in RETURN_STATUSES)
        metrics[platform] = {
            "orders": total_orders,
            "completed_orders": completed_orders,
            "revenue": revenue,
            "completed_revenue": completed_revenue,
            "aov": revenue / total_orders if total_orders else 0,
            "completed_rate": completed_orders / total_orders if total_orders else 0,
            "cancel_rate": cancelled_orders / total_orders if total_orders else 0,
            "return_rate": returned_orders / total_orders if total_orders else 0,
            "at_risk_revenue": at_risk_revenue,
            "cancelled_revenue": cancelled_revenue,
            "status_counts": status_counts,
        }

    total_orders = sum(value["orders"] for value in metrics.values())
    total_revenue = sum(value["revenue"] for value in metrics.values())
    overall = {
        "orders": total_orders,
        "completed_orders": sum(value["completed_orders"] for value in metrics.values()),
        "revenue": total_revenue,
        "completed_revenue": sum(value["completed_revenue"] for value in metrics.values()),
        "aov": total_revenue / total_orders if total_orders else 0,
        "completed_rate": (
            sum(value["completed_orders"] for value in metrics.values()) / total_orders
            if total_orders
            else 0
        ),
        "cancel_rate": (
            sum(value["status_counts"]["Cancelled"] for value in metrics.values()) / total_orders
            if total_orders
            else 0
        ),
        "return_rate": (
            sum(sum(value["status_counts"][status] for status in RETURN_STATUSES) for value in metrics.values())
            / total_orders
            if total_orders
            else 0
        ),
    }
    for value in metrics.values():
        value["order_share"] = value["orders"] / total_orders if total_orders else 0
        value["revenue_share"] = value["revenue"] / total_revenue if total_revenue else 0
    return metrics, overall


def replace_sheet(workbook, name: str, index: int | None = None):
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    return workbook.create_sheet(name, index)


def add_card(ws, cell_range: str, title: str, value: Any, number_format: str, fill: str):
    ws.merge_cells(cell_range)
    start = cell_range.split(":")[0]
    cell = ws[start]
    cell.value = f"{title}\n{value}"
    cell.font = Font(size=13, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.number_format = number_format


def style_table(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    thin = Side(style="thin", color="B7C9D6")
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")


def build_analysis_data(workbook, headers: list[str], records: list[dict[str, Any]]):
    ws = replace_sheet(workbook, DATA_SHEET)
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18
    ws.column_dimensions["K"].width = 48
    for cell in ws["E"][1:]:
        cell.number_format = '#,##0.00'
    ws.sheet_state = "hidden"


def build_summary(workbook, records, metrics, overall):
    ws = replace_sheet(workbook, SUMMARY_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    for column in range(1, 17):
        ws.column_dimensions[get_column_letter(column)].width = 13
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    ws.merge_cells("A1:P1")
    ws["A1"] = "IMURA Executive Summary"
    ws["A1"].font = Font(size=22, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    dates = [as_datetime(row.get("วันที่สั่งซื้อ")) for row in records]
    dates = [value for value in dates if value]
    date_min, date_max = min(dates), max(dates)
    ws.merge_cells("A2:P2")
    ws["A2"] = (
        f"ช่วงข้อมูล {date_min:%d %b %Y} - {date_max:%d %b %Y}  |  "
        f"อัปเดตแดชบอร์ด {datetime.now():%d %b %Y %H:%M}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="44546A")

    cards = [
        ("A4:D6", "จำนวนออเดอร์ทั้งหมด", f"{overall['orders']:,.0f}", '#,##0', BLUE),
        ("E4:H6", "จำนวนออเดอร์สำเร็จ", f"{overall['completed_orders']:,.0f}", '#,##0', GREEN),
        ("I4:L6", "รายได้ทั้งหมด", f"{overall['revenue']:,.2f} บาท", '#,##0.00', ORANGE),
        ("M4:P6", "รายได้จากออเดอร์สำเร็จ", f"{overall['completed_revenue']:,.2f} บาท", '#,##0.00', GREEN),
        ("A8:D10", "มูลค่าเฉลี่ยต่อออเดอร์", f"{overall['aov']:,.2f} บาท", '#,##0.00', BLUE),
        ("E8:H10", "อัตราออเดอร์สำเร็จ", f"{overall['completed_rate']:.1%}", '0.0%', GREEN),
        ("I8:L10", "อัตราการยกเลิก", f"{overall['cancel_rate']:.1%}", '0.0%', RED),
        ("M8:P10", "อัตราการคืนสินค้า", f"{overall['return_rate']:.1%}", '0.0%', ORANGE),
    ]
    for card in cards:
        add_card(ws, *card)
    for row in (4, 5, 6, 8, 9, 10):
        ws.row_dimensions[row].height = 22

    ws.merge_cells("A12:P12")
    ws["A12"] = "เปรียบเทียบประสิทธิภาพแต่ละแพลตฟอร์ม"
    ws["A12"].font = Font(size=15, bold=True, color=WHITE)
    ws["A12"].fill = PatternFill("solid", fgColor=BLUE)

    comparison_headers = [
        "แพลตฟอร์ม", "จำนวนออเดอร์", "ออเดอร์สำเร็จ", "รายได้", "รายได้ออเดอร์สำเร็จ",
        "มูลค่าเฉลี่ยต่อออเดอร์", "สัดส่วนออเดอร์", "สัดส่วนรายได้", "อัตราสำเร็จ", "อัตรายกเลิก",
        "อัตราคืนสินค้า", "รายได้ที่ยังมีความเสี่ยง",
    ]
    for column, header in enumerate(comparison_headers, 1):
        ws.cell(13, column, header)
        ws.cell(13, column).font = Font(bold=True, color=WHITE)
        ws.cell(13, column).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(13, column).alignment = Alignment(horizontal="center", wrap_text=True)

    for row_index, platform in enumerate(PLATFORMS, 14):
        value = metrics[platform]
        row = [
            platform, value["orders"], value["completed_orders"], value["revenue"],
            value["completed_revenue"], value["aov"], value["order_share"],
            value["revenue_share"], value["completed_rate"], value["cancel_rate"],
            value["return_rate"], value["at_risk_revenue"],
        ]
        for column, item in enumerate(row, 1):
            ws.cell(row_index, column, item)
        ws.cell(row_index, 1).font = Font(bold=True)

    for row in range(14, 17):
        for column in (4, 5, 6, 12):
            ws.cell(row, column).number_format = '#,##0.00'
        for column in (7, 8, 9, 10, 11):
            ws.cell(row, column).number_format = '0.0%'
    style_table(ws, 13, 16, 1, 12)

    winners = [
        ("แพลตฟอร์มที่มีออเดอร์สูงสุด", max(PLATFORMS, key=lambda p: metrics[p]["orders"]), "orders", "#,##0"),
        ("แพลตฟอร์มที่มีรายได้สูงสุด", max(PLATFORMS, key=lambda p: metrics[p]["revenue"]), "revenue", "#,##0.00"),
        ("แพลตฟอร์มที่มีมูลค่าต่อออเดอร์สูงสุด", max(PLATFORMS, key=lambda p: metrics[p]["aov"]), "aov", "#,##0.00"),
        ("แพลตฟอร์มที่มีอัตราสำเร็จสูงสุด", max(PLATFORMS, key=lambda p: metrics[p]["completed_rate"]), "completed_rate", "0.0%"),
    ]
    ws.merge_cells("A18:P18")
    ws["A18"] = "แพลตฟอร์มผู้นำในแต่ละด้าน"
    ws["A18"].font = Font(size=15, bold=True, color=WHITE)
    ws["A18"].fill = PatternFill("solid", fgColor=GREEN)
    leader_ranges = ["A19:D22", "E19:H22", "I19:L22", "M19:P22"]
    for item, area in zip(winners, leader_ranges):
        title, platform, metric, fmt = item
        value = metrics[platform][metric]
        display = f"{value:.1%}" if "%" in fmt else f"{value:,.2f}" if ".00" in fmt else f"{value:,.0f}"
        add_card(ws, area, title, f"{platform}\n{display}", fmt, GREEN)

    # Chart source tables on the right side of the dashboard.
    source_col = 18
    ws.cell(2, source_col, "แพลตฟอร์ม")
    ws.cell(2, source_col + 1, "รายได้")
    ws.cell(2, source_col + 2, "จำนวนออเดอร์")
    for index, platform in enumerate(PLATFORMS, 3):
        ws.cell(index, source_col, platform)
        ws.cell(index, source_col + 1, metrics[platform]["revenue"])
        ws.cell(index, source_col + 2, metrics[platform]["orders"])

    status_groups = ["สำเร็จ", "ยกเลิก", "กำลังจัดส่ง / ดำเนินการ", "คืนสินค้า / อื่น ๆ"]
    ws.cell(8, source_col, "แพลตฟอร์ม")
    for index, status in enumerate(status_groups, source_col + 1):
        ws.cell(8, index, status)
    for row_index, platform in enumerate(PLATFORMS, 9):
        counts = metrics[platform]["status_counts"]
        ws.cell(row_index, source_col, platform)
        ws.cell(row_index, source_col + 1, counts["Completed"])
        ws.cell(row_index, source_col + 2, counts["Cancelled"])
        ws.cell(
            row_index,
            source_col + 3,
            sum(counts[status] for status in AT_RISK_STATUSES),
        )
        ws.cell(
            row_index,
            source_col + 4,
            sum(counts.values())
            - counts["Completed"]
            - counts["Cancelled"]
            - sum(counts[status] for status in AT_RISK_STATUSES),
        )

    revenue_chart = DoughnutChart()
    revenue_chart.title = "สัดส่วนรายได้"
    revenue_chart.height = 7
    revenue_chart.width = 9
    revenue_chart.add_data(Reference(ws, min_col=source_col + 1, min_row=2, max_row=5), titles_from_data=True)
    revenue_chart.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=5))
    revenue_chart.dataLabels = DataLabelList()
    revenue_chart.dataLabels.showPercent = True
    revenue_chart.holeSize = 55
    revenue_chart.visible_cells_only = False
    ws.add_chart(revenue_chart, "A24")

    orders_chart = DoughnutChart()
    orders_chart.title = "สัดส่วนจำนวนออเดอร์"
    orders_chart.height = 7
    orders_chart.width = 9
    orders_chart.add_data(Reference(ws, min_col=source_col + 2, min_row=2, max_row=5), titles_from_data=True)
    orders_chart.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=5))
    orders_chart.dataLabels = DataLabelList()
    orders_chart.dataLabels.showPercent = True
    orders_chart.holeSize = 55
    orders_chart.visible_cells_only = False
    ws.add_chart(orders_chart, "I24")

    revenue_bar = BarChart()
    revenue_bar.type = "col"
    revenue_bar.style = 10
    revenue_bar.title = "รายได้แยกตามแพลตฟอร์ม"
    revenue_bar.y_axis.title = "รายได้ (บาท)"
    revenue_bar.height = 7
    revenue_bar.width = 9
    revenue_bar.add_data(Reference(ws, min_col=source_col + 1, min_row=2, max_row=5), titles_from_data=True)
    revenue_bar.set_categories(Reference(ws, min_col=source_col, min_row=3, max_row=5))
    revenue_bar.visible_cells_only = False
    ws.add_chart(revenue_bar, "A39")

    status_chart = BarChart()
    status_chart.type = "bar"
    status_chart.grouping = "stacked"
    status_chart.overlap = 100
    status_chart.style = 12
    status_chart.title = "คุณภาพยอดขายแยกตามสถานะ"
    status_chart.height = 7
    status_chart.width = 9
    status_chart.add_data(
        Reference(ws, min_col=source_col + 1, max_col=source_col + 4, min_row=8, max_row=11),
        titles_from_data=True,
    )
    status_chart.set_categories(Reference(ws, min_col=source_col, min_row=9, max_row=11))
    status_chart.visible_cells_only = False
    ws.add_chart(status_chart, "I39")

    ws.merge_cells("A54:P54")
    ws["A54"] = "สรุปข้อมูลเชิงลึก"
    ws["A54"].font = Font(size=15, bold=True, color=WHITE)
    ws["A54"].fill = PatternFill("solid", fgColor=ORANGE)

    revenue_leader = max(PLATFORMS, key=lambda p: metrics[p]["revenue"])
    order_leader = max(PLATFORMS, key=lambda p: metrics[p]["orders"])
    aov_leader = max(PLATFORMS, key=lambda p: metrics[p]["aov"])
    cancel_high = max(PLATFORMS, key=lambda p: metrics[p]["cancel_rate"])
    completed_leader = max(PLATFORMS, key=lambda p: metrics[p]["completed_rate"])
    insights = [
        f"{revenue_leader} สร้าง Revenue สูงที่สุด {metrics[revenue_leader]['revenue']:,.2f} บาท คิดเป็น {metrics[revenue_leader]['revenue_share']:.1%} ของรายได้รวม",
        f"{order_leader} มี Orders สูงที่สุด {metrics[order_leader]['orders']:,.0f} Orders คิดเป็น {metrics[order_leader]['order_share']:.1%} ของ Orders ทั้งหมด",
        f"{aov_leader} มี Average Order Value สูงที่สุด {metrics[aov_leader]['aov']:,.2f} บาทต่อ Order",
        f"{completed_leader} มี Completed Rate สูงที่สุด {metrics[completed_leader]['completed_rate']:.1%}",
        f"{cancel_high} มี Cancellation Rate สูงที่สุด {metrics[cancel_high]['cancel_rate']:.1%} ควรตรวจสอบสาเหตุการยกเลิก",
    ]
    for index, insight in enumerate(insights, 55):
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=16)
        ws.cell(index, 1, f"• {insight}")
        ws.cell(index, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(index, 1).fill = PatternFill("solid", fgColor=LIGHT_ORANGE if index % 2 else GRAY)
        ws.row_dimensions[index].height = 28

    for column in range(source_col, source_col + 6):
        ws.column_dimensions[get_column_letter(column)].hidden = True


def main():
    headers, records = read_clean_all()
    metrics, overall = calculate_metrics(records)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ANALYSIS_FILE.with_name(f"{ANALYSIS_FILE.stem}.backup_{stamp}{ANALYSIS_FILE.suffix}")
    shutil.copy2(ANALYSIS_FILE, backup)

    workbook = load_workbook(ANALYSIS_FILE)
    build_analysis_data(workbook, headers, records)
    build_summary(workbook, records, metrics, overall)
    if "Sheet1" in workbook.sheetnames and workbook["Sheet1"].max_row == 1:
        workbook.remove(workbook["Sheet1"])
    workbook.save(ANALYSIS_FILE)

    print(f"Updated: {ANALYSIS_FILE}")
    print(f"Backup: {backup}")
    print(f"Orders: {overall['orders']}")
    print(f"Revenue: {overall['revenue']:.2f}")


if __name__ == "__main__":
    main()
