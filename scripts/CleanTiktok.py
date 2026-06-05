from collections import OrderedDict
from datetime import datetime
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RAW_FILE_PATH = Path(r"C:\Education\PlatformSale\scripts\RealRawTiktok-2026-06-05-14_17.xlsx")
RAW_SHEET_NAME = "OrderSKUList"
# OUTPUT_JSON_PATH = Path(r"C:\Education\PlatformSale\scripts\RealRawTiktok-2026-06-05-14_17.json")
# GROUPED_JSON_PATH = Path(r"C:\Education\PlatformSale\scripts\RealRawTiktok-2026-06-05-14_17_grouped_by_order.json")
MASTER_FILE_PATH = Path(r"C:\Education\PlatformSale\scripts\ImuraMasterTiktok.xlsx")

MASTER_ORDERS_SHEET = "Master_Orders"
MASTER_ITEMS_SHEET = "Master_Items"
CLEAN_TIKTOK_SHEET = "Clean_Tiktok"
IMPORT_LOG_SHEET = "Import_Log"

DATE_FIELDS = {
    "Created Time",
    "Paid Time",
    "RTS Time",
    "Shipped Time",
    "Delivered Time",
    "Cancelled Time",
}


def parse_tiktok_datetime(value):
    if value in (None, ""):
        return None

    text = str(value).strip()
    if text in ("", "-"):
        return None

    formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        #"%Y-%m-%d %H:%M:%S",
    )
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format).isoformat(sep=" ")
        except ValueError:
            continue

    return text


def to_text(value):
    if value is None:
        return ""
    return str(value)


def to_decimal(value):
    if value in (None, "", "-"):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def decimal_to_json_value(value):
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def parse_iso_datetime(value):
    if value in (None, ""):
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def first_item_value(order, field_name):
    for item in order.get("Items", []):
        value = item.get(field_name, "")
        if value not in (None, ""):
            return value
    return ""


def normalize_status(order):
    return_type = str(order.get("Cancelation/Return Type", "") or "").strip()
    status = str(order.get("Order Status", "") or "").strip()
    substatus = str(order.get("Order Substatus", "") or "").strip()

    if return_type == "Return/Refund":
        return "Returned"
    if status == "ยกเลิกแล้ว":
        return "Cancelled"
    if status == "จัดส่งแล้ว":
        return "Completed"
    if status == "เสร็จสมบูรณ์":
        return "Completed"
    return status


def group_records_by_order(records):
    grouped_orders = OrderedDict()

    order_level_fields = [
        "Order ID",
        "Order Status",
        "Order Substatus",
        "Cancelation/Return Type",
        "Normal or Pre-order",
        "Order Amount",
        "Order Refund Amount",
        "Created Time",
        "Paid Time",
        "RTS Time",
        "Shipped Time",
        "Delivered Time",
        "Cancelled Time",
        "Cancel By",
        "Cancel Reason",
        "Fulfillment Type",
        "Warehouse Name",
        "Tracking ID",
        "Delivery Option",
        "Shipping Provider Name",
        "Buyer Message",
        "Buyer Username",
        "Recipient",
        "Phone #",
        "Zipcode",
        "Country",
        "Province",
        "District",
        "Districts",
        "Detail Address",
        "Additional address information",
        "Payment Method",
        "Package ID",
        "Seller Note",
        "Checked Status",
        "Checked Marked by",
        "Request Tax Invoice",
        "Tax Info - Buyer Tax ID",
        "Tax Info - Type",
        "Tax Info - Full Name of Buyer",
        "Tax Info - Email",
        "Tax Info - Phone Number",
        "Tax Info - Registered Address",
        "Tax Info - Address Type",
    ]

    item_level_fields = [
        "SKU ID",
        "Seller SKU",
        "Product Name",
        "Variation",
        "Quantity",
        "Sku Quantity of return",
        "SKU Unit Original Price",
        "SKU Subtotal Before Discount",
        "SKU Platform Discount",
        "SKU Seller Discount",
        "SKU Subtotal After Discount",
        "Shipping Fee After Discount",
        "Original Shipping Fee",
        "Shipping Fee Seller Discount",
        "Shipping Fee Platform Discount",
        "Payment platform discount",
        "Taxes",
        "Weight(kg)",
        "Product Category",
    ]

    for record in records:
        order_id = record.get("Order ID", "").strip()
        if not order_id:
            continue

        if order_id not in grouped_orders:
            grouped_orders[order_id] = {
                field_name: record.get(field_name, "")
                for field_name in order_level_fields
                if field_name in record
            }
            grouped_orders[order_id]["Item_Line_Count"] = 0
            grouped_orders[order_id]["Total_Quantity"] = 0
            grouped_orders[order_id]["Items"] = []

        item = {
            field_name: record.get(field_name, "")
            for field_name in item_level_fields
            if field_name in record
        }

        grouped_orders[order_id]["Items"].append(item)
        grouped_orders[order_id]["Item_Line_Count"] += 1

        total_quantity = to_decimal(grouped_orders[order_id]["Total_Quantity"])
        total_quantity += to_decimal(record.get("Quantity"))
        grouped_orders[order_id]["Total_Quantity"] = decimal_to_json_value(total_quantity)

    return list(grouped_orders.values())


def get_order_headers():
    return [
        "Order ID",
        "Order Status",
        "Order Substatus",
        "Cancelation/Return Type",
        "Normal or Pre-order",
        "Order Amount",
        "Order Refund Amount",
        "Created Time",
        "Paid Time",
        "RTS Time",
        "Shipped Time",
        "Delivered Time",
        "Cancelled Time",
        "Cancel By",
        "Cancel Reason",
        "Fulfillment Type",
        "Warehouse Name",
        "Tracking ID",
        "Delivery Option",
        "Shipping Provider Name",
        "Buyer Message",
        "Buyer Username",
        "Recipient",
        "Phone #",
        "Zipcode",
        "Country",
        "Province",
        "District",
        "Districts",
        "Detail Address",
        "Additional address information",
        "Payment Method",
        "Package ID",
        "Seller Note",
        "Checked Status",
        "Checked Marked by",
        "Request Tax Invoice",
        "Tax Info - Buyer Tax ID",
        "Tax Info - Type",
        "Tax Info - Full Name of Buyer",
        "Tax Info - Email",
        "Tax Info - Phone Number",
        "Tax Info - Registered Address",
        "Tax Info - Address Type",
        "Item_Line_Count",
        "Total_Quantity",
        "ImportFileName",
        "LastImportedAt",
    ]


def get_item_headers():
    return [
        "Order ID",
        "Item_No",
        "SKU ID",
        "Seller SKU",
        "Product Name",
        "Variation",
        "Quantity",
        "Sku Quantity of return",
        "SKU Unit Original Price",
        "SKU Subtotal Before Discount",
        "SKU Platform Discount",
        "SKU Seller Discount",
        "SKU Subtotal After Discount",
        "Shipping Fee After Discount",
        "Original Shipping Fee",
        "Shipping Fee Seller Discount",
        "Shipping Fee Platform Discount",
        "Payment platform discount",
        "Taxes",
        "Weight(kg)",
        "Product Category",
        "ImportFileName",
        "LastImportedAt",
    ]


def get_clean_tiktok_headers():
    return [
        "หมายเลขคำสั่งซื้อออนไลน์",
        "วันที่สั่งซื้อ",
        "แพลตฟอร์ม",
        "SKU",
        "จำนวนเงินที่ควรได้รับ",
        "จังหวัด",
        "Order_Status_Clean",
        "Purchase_Hour",
        "Day_of_Week",
        "Month_Year",
        "Product_NAME",
        "ตัวเลือก",
        "Item_Line_Count",
        "Total_Quantity",
    ]


def load_existing_master_orders(workbook):
    if MASTER_ORDERS_SHEET not in workbook.sheetnames:
        return OrderedDict()

    worksheet = workbook[MASTER_ORDERS_SHEET]
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return OrderedDict()

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    existing_orders = OrderedDict()
    for row in rows[1:]:
        record = {
            header: "" if index >= len(row) or row[index] is None else row[index]
            for index, header in enumerate(headers)
            if header
        }
        order_id = str(record.get("Order ID", "")).strip()
        if order_id:
            existing_orders[order_id] = record

    return existing_orders


def open_or_create_master_workbook():
    if MASTER_FILE_PATH.exists():
        return load_workbook(MASTER_FILE_PATH)
    return Workbook()


def replace_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name)


def style_worksheet(worksheet, freeze_cell="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = freeze_cell
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, cells in enumerate(worksheet.iter_cols(min_row=1, max_row=min(worksheet.max_row, 100)), start=1):
        max_length = 12
        for cell in cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)


def write_master_orders(workbook, orders):
    worksheet = replace_sheet(workbook, MASTER_ORDERS_SHEET)
    headers = get_order_headers()
    worksheet.append(headers)
    for order in orders:
        worksheet.append([order.get(header, "") for header in headers])
    style_worksheet(worksheet)


def write_master_items(workbook, orders):
    worksheet = replace_sheet(workbook, MASTER_ITEMS_SHEET)
    headers = get_item_headers()
    worksheet.append(headers)
    for order in orders:
        order_id = order.get("Order ID", "")
        for item_no, item in enumerate(order.get("Items", []), start=1):
            row = []
            for header in headers:
                if header == "Order ID":
                    row.append(order_id)
                elif header == "Item_No":
                    row.append(item_no)
                elif header in ("ImportFileName", "LastImportedAt"):
                    row.append(order.get(header, ""))
                else:
                    row.append(item.get(header, ""))
            worksheet.append(row)
    style_worksheet(worksheet)


def write_clean_tiktok(workbook, orders):
    worksheet = replace_sheet(workbook, CLEAN_TIKTOK_SHEET)
    worksheet.append(get_clean_tiktok_headers())

    for order in orders:
        created_at = parse_iso_datetime(order.get("Created Time"))
        month_year = datetime(created_at.year, created_at.month, 1) if created_at else ""

        worksheet.append(
            [
                order.get("Order ID", ""),
                created_at or order.get("Created Time", ""),
                "TikTok",
                first_item_value(order, "Seller SKU"),
                float(to_decimal(order.get("Order Amount"))),
                order.get("Province", ""),
                normalize_status(order),
                created_at.hour if created_at else "",
                created_at.strftime("%A") if created_at else "",
                month_year,
                first_item_value(order, "Product Name"),
                first_item_value(order, "Variation"),
                order.get("Item_Line_Count", 0),
                order.get("Total_Quantity", 0),
            ]
        )

    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].number_format = "#,##0.00"
    for row in worksheet.iter_rows(min_row=2, min_col=10, max_col=10):
        row[0].number_format = "yyyy-mm-dd"

    style_worksheet(worksheet)


def append_import_log(workbook, summary):
    if IMPORT_LOG_SHEET not in workbook.sheetnames:
        worksheet = workbook.create_sheet(IMPORT_LOG_SHEET)
        worksheet.append(
            [
                "ImportedAt",
                "ImportFileName",
                "RawRows",
                "IncomingOrders",
                "InsertedOrders",
                "UpdatedOrders",
                "TotalMasterOrders",
            ]
        )
    else:
        worksheet = workbook[IMPORT_LOG_SHEET]

    worksheet.append(
        [
            summary["imported_at"],
            summary["import_file_name"],
            summary["raw_rows"],
            summary["incoming_orders"],
            summary["inserted_orders"],
            summary["updated_orders"],
            summary["total_master_orders"],
        ]
    )
    style_worksheet(worksheet)


def upsert_master_dataset(grouped_records):
    imported_at = datetime.now().isoformat(sep=" ", timespec="seconds")
    workbook = open_or_create_master_workbook()
    existing_orders = load_existing_master_orders(workbook)

    inserted_orders = 0
    updated_orders = 0

    for order in grouped_records:
        order_id = str(order.get("Order ID", "")).strip()
        if not order_id:
            continue

        if order_id in existing_orders:
            updated_orders += 1
        else:
            inserted_orders += 1

        order["ImportFileName"] = RAW_FILE_PATH.name
        order["LastImportedAt"] = imported_at
        existing_orders[order_id] = order

    master_orders = list(existing_orders.values())
    write_master_orders(workbook, master_orders)
    write_master_items(workbook, master_orders)
    write_clean_tiktok(workbook, master_orders)
    append_import_log(
        workbook,
        {
            "imported_at": imported_at,
            "import_file_name": RAW_FILE_PATH.name,
            "raw_rows": sum(order.get("Item_Line_Count", 0) for order in grouped_records),
            "incoming_orders": len(grouped_records),
            "inserted_orders": inserted_orders,
            "updated_orders": updated_orders,
            "total_master_orders": len(master_orders),
        },
    )

    if "Sheet1" in workbook.sheetnames and workbook["Sheet1"].max_row == 1 and workbook["Sheet1"].max_column == 1:
        del workbook["Sheet1"]

    workbook.save(MASTER_FILE_PATH)
    workbook.close()

    return {
        "inserted_orders": inserted_orders,
        "updated_orders": updated_orders,
        "total_master_orders": len(master_orders),
    }


def main():
    if not RAW_FILE_PATH.exists():
        raise FileNotFoundError(RAW_FILE_PATH)

    workbook = load_workbook(RAW_FILE_PATH, read_only=True, data_only=True)

    if RAW_SHEET_NAME not in workbook.sheetnames:
        available_sheets = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{RAW_SHEET_NAME}' not found. Available sheets: {available_sheets}")

    worksheet = workbook[RAW_SHEET_NAME]
    headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {
        str(header).strip(): index
        for index, header in enumerate(headers)
        if header not in (None, "")
    }

    print(f"File: {RAW_FILE_PATH}")
    print(f"Sheet: {RAW_SHEET_NAME}")
    print(f"Rows: {worksheet.max_row:,}")
    print(f"Columns: {worksheet.max_column:,}")
    print(f"Header count: {len(header_index):,}")

    print("\nDate field samples:")
    for field_name in DATE_FIELDS:
        if field_name not in header_index:
            print(f"- {field_name}: missing")
            continue

        column_index = header_index[field_name]
        samples = []
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            value = row[column_index] if column_index < len(row) else None
            if value not in (None, ""):
                samples.append(value)
            if len(samples) >= 3:
                break

        print(f"- {field_name}: {samples}")

    print("\nFirst 5 data rows:")
    preview_fields = ["Order ID", "Order Status", "Order Substatus", "Created Time", "Order Amount"]
    for row_number, row in enumerate(worksheet.iter_rows(min_row=3, max_row=7, values_only=True), start=3):
        preview = {}
        for field_name in preview_fields:
            column_index = header_index.get(field_name)
            preview[field_name] = row[column_index] if column_index is not None and column_index < len(row) else None
        print(f"Row {row_number}: {preview}")

    records = []
    ordered_headers = [str(header).strip() for header in headers if header not in (None, "")]

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        record = {}
        has_value = False

        for field_name in ordered_headers:
            column_index = header_index[field_name]
            value = row[column_index] if column_index < len(row) else None

            if value not in (None, ""):
                has_value = True

            if field_name in DATE_FIELDS:
                record[field_name] = parse_tiktok_datetime(value)
            else:
                record[field_name] = to_text(value)

        if has_value:
            records.append(record)

    grouped_records = group_records_by_order(records)

    print(f"\nJSON rows: {len(records):,}")

    master_summary = upsert_master_dataset(grouped_records)
    print("\nMaster update:")
    print(f"Inserted orders: {master_summary['inserted_orders']:,}")
    print(f"Updated orders: {master_summary['updated_orders']:,}")
    print(f"Total master orders: {master_summary['total_master_orders']:,}")
    print(f"Master file: {MASTER_FILE_PATH}")

    workbook.close()


if __name__ == "__main__":
    main()
