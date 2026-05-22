from __future__ import annotations

from collections import Counter, OrderedDict
from dataclasses import dataclass
from statistics import mean, stdev
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from shopee_order import ShopeeOrder


CANCELLED_STATUS = "ยกเลิกแล้ว"


@dataclass(frozen=True)
class FeeSummary:
    order_count: int
    min_percent: float
    mean_percent: float
    std_percent: float
    max_percent: float
    rounded_buckets: OrderedDict[int, int]


@dataclass(frozen=True)
class FeeSheetResult:
    sheet_name: str
    row_count: int
    summary: FeeSummary


def to_float(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    return float(str(value).replace(",", ""))


def group_active_orders(orders: list[ShopeeOrder]) -> OrderedDict[str, list[ShopeeOrder]]:
    grouped: OrderedDict[str, list[ShopeeOrder]] = OrderedDict()
    for order in orders:
        if str(order.order_status).strip() == CANCELLED_STATUS:
            continue
        if to_float(order.returned_quantity) > 0:
            continue
        grouped.setdefault(str(order.order_id), []).append(order)
    return grouped


def calculate_fee_summary(percent_values: list[float]) -> FeeSummary:
    rounded_counts = Counter(round(value * 100) for value in percent_values)
    rounded_buckets = OrderedDict(
        sorted(rounded_counts.items(), key=lambda item: (-item[1], item[0]))
    )

    if not percent_values:
        return FeeSummary(0, 0.0, 0.0, 0.0, 0.0, rounded_buckets)

    return FeeSummary(
        order_count=len(percent_values),
        min_percent=min(percent_values),
        mean_percent=mean(percent_values),
        std_percent=stdev(percent_values) if len(percent_values) > 1 else 0.0,
        max_percent=max(percent_values),
        rounded_buckets=rounded_buckets,
    )


def style_report_sheet(
    ws,
    *,
    title: str,
    subtitle: str,
    headers: list[str],
    row_count: int,
    percent_field: str,
    money_fields: set[str],
    field_names: list[str],
    table_name: str,
) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2F3")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="44546A")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_index, value=header)
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = header_fill
        cell.border = Border(bottom=thin_gray)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(5, 5 + row_count):
        for col_index, field_name in enumerate(field_names, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.alignment = Alignment(vertical="top")
            if field_name in money_fields:
                cell.number_format = "#,##0.00"
            elif field_name == percent_field:
                cell.number_format = "0.00%"
            elif field_name == "item_count":
                cell.number_format = "0"

    end_row = max(4 + row_count, 5)
    end_col = len(headers)
    table = Table(displayName=table_name, ref=f"A4:{get_column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A5"

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) for cell in column_cells[:80] if cell.value is not None)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 46)


def add_summary_sheet(wb, sheet_results: list[FeeSheetResult], *, raw_row_count: int, active_order_count: int) -> None:
    ws = wb.create_sheet("Summary", 0)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2F3")

    ws["A1"] = "Shopee Fee Summary"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    ws["A2"] = "สรุปจากข้อมูลระดับ order หลังกรองคำสั่งซื้อยกเลิกและรายการคืนสินค้าออกแล้ว"
    ws["A2"].font = Font(italic=True, color="44546A")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    ws.append([])
    ws.append(["Metric", "Value"])
    ws.append(["Raw item rows", raw_row_count])
    ws.append(["Active order count", active_order_count])

    summary_start = 8
    summary_headers = ["Fee Sheet", "Count", "Min", "Mean", "Std", "Max"]
    for col_index, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_start, column=col_index, value=header)
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = header_fill
        cell.border = Border(bottom=thin_gray)
        cell.alignment = Alignment(horizontal="center")

    for row_index, result in enumerate(sheet_results, start=summary_start + 1):
        summary = result.summary
        values = [
            result.sheet_name,
            summary.order_count,
            summary.min_percent,
            summary.mean_percent,
            summary.std_percent,
            summary.max_percent,
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if col_index >= 3:
                cell.number_format = "0.00%"

    bucket_start = summary_start + len(sheet_results) + 4
    ws.cell(row=bucket_start, column=1, value="Rounded Percent Buckets")
    ws.cell(row=bucket_start, column=1).font = Font(bold=True, color="1F4E78")

    row_index = bucket_start + 2
    for result in sheet_results:
        ws.cell(row=row_index, column=1, value=result.sheet_name)
        ws.cell(row=row_index, column=1).font = Font(bold=True, color="1F4E78")
        ws.cell(row=row_index + 1, column=1, value="Bucket")
        ws.cell(row=row_index + 1, column=2, value="Orders")
        for col in range(1, 3):
            cell = ws.cell(row=row_index + 1, column=col)
            cell.font = Font(bold=True, color="1F4E78")
            cell.fill = header_fill
            cell.border = Border(bottom=thin_gray)
            cell.alignment = Alignment(horizontal="center")

        for offset, (bucket, count) in enumerate(result.summary.rounded_buckets.items(), start=2):
            ws.cell(row=row_index + offset, column=1, value=f"{bucket}%")
            ws.cell(row=row_index + offset, column=2, value=count)
        row_index += len(result.summary.rounded_buckets) + 4

    for column in range(1, 7):
        ws.column_dimensions[get_column_letter(column)].width = 22
