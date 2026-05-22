from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


path = Path(__file__).with_name("DATABASE_TODO.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb["Todo"]
print(ws.title, ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=4, max_row=6, values_only=True):
    print(row)
