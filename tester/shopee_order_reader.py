from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from shopee_order import ShopeeOrder


DEFAULT_WORKBOOK_PATH = Path(__file__).with_name("Commission_MML.xlsx")
SHOPEE_ORDER_SHEET_NAME = "Shopee คำสั่งซื้อ"


def _normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _validate_header(header_row: tuple[Any, ...]) -> None:
    actual_columns = tuple(_normalize_header(value) for value in header_row[: len(ShopeeOrder.SOURCE_COLUMNS)])
    expected_columns = ShopeeOrder.SOURCE_COLUMNS

    if actual_columns == expected_columns:
        return

    mismatches = [
        (index + 1, expected, actual)
        for index, (expected, actual) in enumerate(zip(expected_columns, actual_columns))
        if expected != actual
    ]
    preview = "\n".join(
        f"column {column}: expected {expected!r}, got {actual!r}"
        for column, expected, actual in mismatches[:10]
    )
    raise ValueError(
        "Shopee order worksheet headers do not match ShopeeOrder.SOURCE_COLUMNS.\n"
        f"{preview}"
    )


def iter_shopee_orders(
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
    *,
    sheet_name: str = SHOPEE_ORDER_SHEET_NAME,
    validate_header: bool = True,
) -> Iterator[ShopeeOrder]:
    """Yield ShopeeOrder objects from the Shopee order worksheet."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]

        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return

        if validate_header:
            _validate_header(header_row)

        for row in rows:
            row_values = tuple(row[: len(ShopeeOrder.SOURCE_COLUMNS)])
            if not any(value not in (None, "") for value in row_values):
                continue
            yield ShopeeOrder.from_excel_row(row_values)
    finally:
        workbook.close()


def read_shopee_orders(
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
    *,
    sheet_name: str = SHOPEE_ORDER_SHEET_NAME,
    validate_header: bool = True,
) -> list[ShopeeOrder]:
    """Read all Shopee orders into a list."""
    return list(
        iter_shopee_orders(
            workbook_path,
            sheet_name=sheet_name,
            validate_header=validate_header,
        )
    )



