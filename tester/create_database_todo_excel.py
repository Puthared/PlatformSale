from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = Path(__file__).with_name("DATABASE_TODO.xlsx")


def build_workbook() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Todo"

    ws["A1"] = "MS-SQL Platform Order Database Todo"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Initial checklist for using SQLAlchemy and Alembic."
    ws["A2"].font = Font(italic=True, color="44546A")
    ws.merge_cells("A2:F2")

    headers = ["ID", "Phase", "Task", "Status", "Command / Output", "Notes"]
    rows = [
        [1, "Setup", "ลง sqlalchemy และ alembic ใน venv", "Not Started", "pip install sqlalchemy alembic", ""],
        [2, "Database Design", "สร้าง Model Database", "Not Started", "", "เริ่มจาก models สำหรับ orders / order_items / platforms"],
    ]

    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for row_index, row_values in enumerate(rows, start=header_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    table_ref = f"A{header_row}:F{header_row + len(rows)}"
    table = Table(displayName="DatabaseTodoTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A5"

    widths = {
        "A": 8,
        "B": 18,
        "C": 36,
        "D": 16,
        "E": 34,
        "F": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    print(build_workbook())
