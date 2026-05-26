import math
import sys
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from os import path

sys.path.append(path.join(path.dirname(__file__), ".."))

from fastapi import APIRouter, Body, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from models import TiktokMaster
from models.modelsDTO import TiktokMaster as tb
from config import database as db
from services.TiktokNormalizer import normalize_tiktok_master


router = APIRouter(
    prefix="/TiktokMaster",
    tags=["TiktokMaster"],
    responses={
        404: {
            "message": "Not Found"
        }
    }
)

TTMaster = TiktokMaster.TiktokMaster

TIKTOK_SHEET_NAME = "Raw_Tiktok"

TIKTOK_HEADER_MAPPING = [
    ("Order ID", "OrderId"),
    ("Order Status", "OrderStatus"),
    ("Order Substatus", "OrderSubstatus"),
    ("Cancelation/Return Type", "CancelationReturnType"),
    ("Normal or Pre-order", "NormalOrPreOrder"),
    ("SKU ID", "SkuId"),
    ("Seller SKU", "SellerSku"),
    ("Product Name", "ProductName"),
    ("Variation", "Variation"),
    ("Quantity", "Quantity"),
    ("Sku Quantity of return", "SkuQuantityOfReturn"),
    ("SKU Unit Original Price", "SkuUnitOriginalPrice"),
    ("SKU Subtotal Before Discount", "SkuSubtotalBeforeDiscount"),
    ("SKU Platform Discount", "SkuPlatformDiscount"),
    ("SKU Seller Discount", "SkuSellerDiscount"),
    ("SKU Subtotal After Discount", "SkuSubtotalAfterDiscount"),
    ("Shipping Fee After Discount", "ShippingFeeAfterDiscount"),
    ("Original Shipping Fee", "OriginalShippingFee"),
    ("Shipping Fee Seller Discount", "ShippingFeeSellerDiscount"),
    ("Shipping Fee Platform Discount", "ShippingFeePlatformDiscount"),
    ("Payment platform discount", "PaymentPlatformDiscount"),
    ("Taxes", "Taxes"),
    ("Order Amount", "OrderAmount"),
    ("Order Refund Amount", "OrderRefundAmount"),
    ("Created Time", "CreatedTime"),
    ("Paid Time", "PaidTime"),
    ("RTS Time", "RtsTime"),
    ("Shipped Time", "ShippedTime"),
    ("Delivered Time", "DeliveredTime"),
    ("Cancelled Time", "CancelledTime"),
    ("Cancel By", "CancelBy"),
    ("Cancel Reason", "CancelReason"),
    ("Fulfillment Type", "FulfillmentType"),
    ("Warehouse Name", "WarehouseName"),
    ("Tracking ID", "TrackingId"),
    ("Delivery Option", "DeliveryOption"),
    ("Shipping Provider Name", "ShippingProviderName"),
    ("Buyer Message", "BuyerMessage"),
    ("Buyer Username", "BuyerUsername"),
    ("Recipient", "Recipient"),
    ("Phone #", "Phone"),
    ("Zipcode", "Zipcode"),
    ("Country", "Country"),
    ("Province", "Province"),
    ("District", "District"),
    ("Districts", "Districts"),
    ("Detail Address", "DetailAddress"),
    ("Additional address information", "AdditionalAddressInformation"),
    ("Payment Method", "PaymentMethod"),
    ("Weight(kg)", "WeightKg"),
    ("Product Category", "ProductCategory"),
    ("Package ID", "PackageId"),
    ("Seller Note", "SellerNote"),
    ("Checked Status", "CheckedStatus"),
    ("Checked Marked by", "CheckedMarkedBy"),
    ("Request Tax Invoice", "RequestTaxInvoice"),
    ("Tax Info - Buyer Tax ID", "TaxInfoBuyerTaxId"),
    ("Tax Info - Type", "TaxInfoType"),
    ("Tax Info - Full Name of Buyer", "TaxInfoFullNameOfBuyer"),
    ("Tax Info - Email", "TaxInfoEmail"),
    ("Tax Info - Phone Number", "TaxInfoPhoneNumber"),
    ("Tax Info - Registered Address", "TaxInfoRegisteredAddress"),
    ("Tax Info - Address Type", "TaxInfoAddressType"),
]


def _model_columns():
    return {column.name for column in TTMaster.__table__.columns}


def _writable_columns():
    blocked_columns = {
        "TiktokMasterId",
        "isDeleted",
        "createdOn",
        "modifiedOn",
    }
    return _model_columns() - blocked_columns


def _payload_to_dict(payload) -> dict:
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=True)
    if hasattr(payload, "dict"):
        return payload.dict(exclude_unset=True)
    return dict(payload)


def _build_model_payload(payload, *, is_create: bool) -> dict:
    payload = _payload_to_dict(payload)
    writable_columns = _writable_columns()
    model_payload = {
        key: value
        for key, value in payload.items()
        if key in writable_columns
    }

    if is_create:
        model_payload["createdBy"] = payload.get("createdBy") or "system"
        model_payload["createdOn"] = datetime.now()
        model_payload["isDeleted"] = False
    else:
        model_payload["modifiedBy"] = payload.get("modifiedBy") or payload.get("createdBy") or "system"
        model_payload["modifiedOn"] = datetime.now()

    return model_payload


def _normalize_header(value):
    return "" if value is None else str(value).strip()


def _cell_to_text(value):
    if value is None:
        return None
    return str(value)


def _normalize_item_key_part(value):
    if value in (None, ""):
        return ""
    return " ".join(str(value).strip().split())


def _get_item_key_value(row, column_name: str):
    if isinstance(row, dict):
        return row.get(column_name)
    return getattr(row, column_name)


def _build_item_key(row):
    return (
        _normalize_item_key_part(_get_item_key_value(row, "OrderId")),
        _normalize_item_key_part(_get_item_key_value(row, "SkuId")),
        _normalize_item_key_part(_get_item_key_value(row, "SellerSku")),
        _normalize_item_key_part(_get_item_key_value(row, "ProductName")),
        _normalize_item_key_part(_get_item_key_value(row, "Variation")),
    )


def _build_occurrence_map(rows):
    item_key_counts = defaultdict(int)
    mapped_rows = {}
    for row in rows:
        base_key = _build_item_key(row)
        item_key_counts[base_key] += 1
        mapped_rows[(*base_key, item_key_counts[base_key])] = row
    return mapped_rows


def _chunks(values, chunk_size=1000):
    for index in range(0, len(values), chunk_size):
        yield values[index:index + chunk_size]


def _validate_tiktok_headers(headers: list[str]):
    expected_headers = [header for header, _ in TIKTOK_HEADER_MAPPING]
    actual_known_headers = headers[:len(expected_headers)]

    if len(headers) < len(expected_headers):
        return [
            {
                "column": index + 1,
                "expected": expected_headers[index],
                "actual": actual_known_headers[index] if index < len(actual_known_headers) else None,
            }
            for index in range(len(expected_headers))
            if index >= len(actual_known_headers) or expected_headers[index] != actual_known_headers[index]
        ]

    return [
        {
            "column": index + 1,
            "expected": expected_header,
            "actual": actual_header,
        }
        for index, (expected_header, actual_header) in enumerate(zip(expected_headers, actual_known_headers))
        if expected_header != actual_header
    ]


@router.get("/GetTiktokMaster")
async def GetTiktokMaster(req: Request, page: int, perpage: int = 1000):
    try:
        TiktokMasterItem = (
            db.session
            .query(TTMaster)
            .filter(TTMaster.isDeleted == False)
            .order_by(TTMaster.TiktokMasterId)
            .all()
        )

        PaginatedRecord = []
        Page = page
        Perpage = perpage
        TotalRecord = TiktokMasterItem.__len__()
        AmountPage = 1 if TotalRecord <= Perpage else math.ceil(TotalRecord / Perpage)

        if Page > AmountPage:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Invaild page pagination.", "data": ""},
            )

        TargetIndex = (Page - 1) * Perpage
        FinishElement = TargetIndex + Perpage
        if Page < AmountPage:
            PaginatedRecord = TiktokMasterItem[TargetIndex:FinishElement]
        else:
            PaginatedRecord = TiktokMasterItem[TargetIndex:]
        PaginatedRecord = [item.as_dict() for item in PaginatedRecord]

        PaginateReturn = {
            "totalRecords": TotalRecord,
            "pageCount": AmountPage,
            "pageNo": Page,
            "pageSize": Perpage,
        }
        return {
            "status": "success",
            "message": "",
            "data": {"data": PaginatedRecord, "pagination": PaginateReturn},
        }

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.get("/GetTiktokMasterById/{tiktokMasterId}")
async def GetTiktokMasterById(req: Request, tiktokMasterId: int):
    try:
        TiktokMasterItem = (
            db.session
            .query(TTMaster)
            .filter(TTMaster.TiktokMasterId == tiktokMasterId)
            .filter(TTMaster.isDeleted == False)
            .first()
        )

        if TiktokMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "TiktokMaster not found.", "data": ""},
            )

        return {"status": "success", "message": "", "data": TiktokMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/ImportTiktokMaster")
async def ImportTiktokMaster(
    req: Request,
    file: UploadFile = File(...),
    createdBy: str = Form("system"),
):
    try:
        file_bytes = await file.read()
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        if TIKTOK_SHEET_NAME not in workbook.sheetnames:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": f"Worksheet '{TIKTOK_SHEET_NAME}' not found.",
                    "data": {"availableSheets": workbook.sheetnames},
                },
            )

        worksheet = workbook[TIKTOK_SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Excel file has no header row.", "data": ""},
            )

        actual_headers = [_normalize_header(value) for value in header_row if value not in (None, "")]
        header_errors = _validate_tiktok_headers(actual_headers)
        if header_errors:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": "TikTok Excel headers are not compatible with TiktokMaster mapping.",
                    "data": {
                        "expectedHeaderCount": len(TIKTOK_HEADER_MAPPING),
                        "actualHeaderCount": len(actual_headers),
                        "headerErrors": header_errors[:30],
                    },
                },
            )

        extra_headers = actual_headers[len(TIKTOK_HEADER_MAPPING):]
        column_names = [column_name for _, column_name in TIKTOK_HEADER_MAPPING]
        now = datetime.now()
        incoming_orders = {}
        skipped_empty_rows = 0
        skipped_missing_order_id = 0

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            known_values = row[:len(column_names)]
            if not any(value not in (None, "") for value in known_values):
                skipped_empty_rows += 1
                continue

            model_payload = {
                column_name: _cell_to_text(value)
                for column_name, value in zip(column_names, known_values)
            }
            order_id = _normalize_item_key_part(model_payload.get("OrderId"))
            if not order_id:
                skipped_missing_order_id += 1
                continue
            incoming_orders.setdefault(order_id, []).append(model_payload)

        order_ids = list(incoming_orders.keys())
        existing_rows = []
        for order_id_chunk in _chunks(order_ids):
            existing_rows.extend(
                db.session
                .query(TTMaster)
                .filter(TTMaster.OrderId.in_(order_id_chunk))
                .order_by(TTMaster.TiktokMasterId)
                .all()
            )

        existing_orders = defaultdict(list)
        for existing_row in existing_rows:
            existing_orders[_normalize_item_key_part(existing_row.OrderId)].append(existing_row)

        records_to_create = []
        inserted_rows = 0
        updated_rows = 0
        soft_deleted_rows = 0
        affected_orders = 0

        for order_id, incoming_rows in incoming_orders.items():
            order_changed = False
            incoming_item_map = _build_occurrence_map(incoming_rows)
            existing_item_map = _build_occurrence_map(existing_orders.get(order_id, []))

            for item_key, incoming_payload in incoming_item_map.items():
                existing_row = existing_item_map.get(item_key)
                if existing_row is None:
                    create_payload = dict(incoming_payload)
                    create_payload["ImportFileName"] = file.filename
                    create_payload["IsNormalized"] = False
                    create_payload["NormalizedOn"] = None
                    create_payload["NormalizeError"] = None
                    create_payload["isDeleted"] = False
                    create_payload["createdBy"] = createdBy
                    create_payload["createdOn"] = now
                    records_to_create.append(TTMaster(**create_payload))
                    inserted_rows += 1
                    order_changed = True
                    continue

                for column_name in column_names:
                    setattr(existing_row, column_name, incoming_payload.get(column_name))
                existing_row.ImportFileName = file.filename
                existing_row.IsNormalized = False
                existing_row.NormalizedOn = None
                existing_row.NormalizeError = None
                existing_row.isDeleted = False
                existing_row.modifiedBy = createdBy
                existing_row.modifiedOn = now
                updated_rows += 1
                order_changed = True

            for item_key, existing_row in existing_item_map.items():
                if item_key in incoming_item_map:
                    continue
                if existing_row.isDeleted is not True:
                    soft_deleted_rows += 1
                existing_row.isDeleted = True
                existing_row.IsNormalized = False
                existing_row.NormalizedOn = None
                existing_row.NormalizeError = None
                existing_row.ImportFileName = file.filename
                existing_row.modifiedBy = createdBy
                existing_row.modifiedOn = now
                order_changed = True

            if order_changed:
                affected_orders += 1

        if records_to_create:
            db.session.add_all(records_to_create)

        db.session.commit()

        return {
            "status": "success",
            "message": "Import TiktokMaster success.",
            "data": {
                "fileName": file.filename,
                "sheetName": TIKTOK_SHEET_NAME,
                "ordersInFile": len(incoming_orders),
                "affectedOrders": affected_orders,
                "insertedRows": inserted_rows,
                "updatedRows": updated_rows,
                "softDeletedRows": soft_deleted_rows,
                "skippedEmptyRows": skipped_empty_rows,
                "skippedMissingOrderIdRows": skipped_missing_order_id,
                "extraHeaders": extra_headers,
            },
        }

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/NormalizeTiktokMaster")
async def NormalizeTiktokMaster(req: Request, body: tb.TiktokMasterNormalizeDTO):
    try:
        result = normalize_tiktok_master(
            limit=body.limit,
            created_by=body.createdBy,
            mode=body.mode,
        )
        return {"status": "success", "message": "Normalize TiktokMaster success.", "data": result}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/CreateTiktokMaster")
async def CreateTiktokMaster(req: Request, body: tb.TiktokMasterCreateDTO):
    try:
        model_payload = _build_model_payload(body, is_create=True)
        TiktokMasterItem = TTMaster(**model_payload)

        db.session.add(TiktokMasterItem)
        db.session.commit()
        db.session.refresh(TiktokMasterItem)

        return {"status": "success", "message": "Create TiktokMaster success.", "data": TiktokMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.put("/UpdateTiktokMaster/{tiktokMasterId}")
async def UpdateTiktokMaster(req: Request, tiktokMasterId: int, body: tb.TiktokMasterUpdateDTO):
    try:
        TiktokMasterItem = (
            db.session
            .query(TTMaster)
            .filter(TTMaster.TiktokMasterId == tiktokMasterId)
            .filter(TTMaster.isDeleted == False)
            .first()
        )

        if TiktokMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "TiktokMaster not found.", "data": ""},
            )

        model_payload = _build_model_payload(body, is_create=False)
        for key, value in model_payload.items():
            setattr(TiktokMasterItem, key, value)

        db.session.commit()
        db.session.refresh(TiktokMasterItem)

        return {"status": "success", "message": "Update TiktokMaster success.", "data": TiktokMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.delete("/DeleteTiktokMaster/{tiktokMasterId}")
async def DeleteTiktokMaster(req: Request, tiktokMasterId: int, payload: dict = Body(default={})):
    try:
        TiktokMasterItem = (
            db.session
            .query(TTMaster)
            .filter(TTMaster.TiktokMasterId == tiktokMasterId)
            .filter(TTMaster.isDeleted == False)
            .first()
        )

        if TiktokMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "TiktokMaster not found.", "data": ""},
            )

        TiktokMasterItem.isDeleted = True
        TiktokMasterItem.modifiedBy = payload.get("modifiedBy") or payload.get("createdBy") or "system"
        TiktokMasterItem.modifiedOn = datetime.now()

        db.session.commit()
        db.session.refresh(TiktokMasterItem)

        return {"status": "success", "message": "Delete TiktokMaster success.", "data": TiktokMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()
