import sys
from dataclasses import dataclass
from io import BytesIO
from os import path
from typing import Callable

sys.path.append(path.join(path.dirname(__file__), ".."))

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from sqlalchemy import func

from config import database as db
from models.ShopeeMaster import ShopeeMaster
from models.TiktokMaster import TiktokMaster
from models.modelsDTO import DataImport as tb
from routes.ShopeeMaster import (
    ImportShopeeMaster,
    SHOPEE_HEADER_MAPPING,
    SHOPEE_SHEET_NAME,
    _validate_shopee_headers,
)
from routes.TikTokMaster import (
    ImportTiktokMaster,
    TIKTOK_HEADER_MAPPING,
    TIKTOK_SHEET_NAME,
    _validate_tiktok_headers,
)
from services.ShopeeNormalizer import normalize_shopee_master
from services.TiktokNormalizer import normalize_tiktok_master


router = APIRouter(
    prefix="/DataImport",
    tags=["DataImport"],
    responses={
        404: {
            "message": "Not Found"
        }
    }
)


@dataclass(frozen=True)
class RawPlatformConfig:
    key: str
    name: str
    master_table: str
    sheet_name: str | None
    model: object | None
    id_column: str | None
    header_mapping: list[tuple[str, str]]
    header_validator: Callable[[list[str]], list[dict]]
    import_handler: Callable | None
    normalizer: Callable | None
    is_supported: bool = True


PLATFORM_CONFIGS = {
    "shopee": RawPlatformConfig(
        key="shopee",
        name="Shopee",
        master_table="ShopeeMaster",
        sheet_name=SHOPEE_SHEET_NAME,
        model=ShopeeMaster,
        id_column="ShopeeMasterId",
        header_mapping=SHOPEE_HEADER_MAPPING,
        header_validator=_validate_shopee_headers,
        import_handler=ImportShopeeMaster,
        normalizer=normalize_shopee_master,
    ),
    "tiktok": RawPlatformConfig(
        key="tiktok",
        name="TikTok",
        master_table="TiktokMaster",
        sheet_name=TIKTOK_SHEET_NAME,
        model=TiktokMaster,
        id_column="TiktokMasterId",
        header_mapping=TIKTOK_HEADER_MAPPING,
        header_validator=_validate_tiktok_headers,
        import_handler=ImportTiktokMaster,
        normalizer=normalize_tiktok_master,
    ),
    "lazada": RawPlatformConfig(
        key="lazada",
        name="Lazada",
        master_table="LazadaMaster",
        sheet_name=None,
        model=None,
        id_column=None,
        header_mapping=[],
        header_validator=lambda headers: [],
        import_handler=None,
        normalizer=None,
        is_supported=False,
    ),
}


def _normalize_platform_key(platform: str) -> str:
    return (platform or "").strip().lower().replace(" ", "")


def _get_platform_config(platform: str) -> RawPlatformConfig:
    platform_key = _normalize_platform_key(platform)
    config = PLATFORM_CONFIGS.get(platform_key)
    if config is None:
        raise ValueError(f"Unsupported platform '{platform}'.")
    return config


def _normalize_header(value) -> str:
    return "" if value is None else str(value).strip()


def _platform_to_dict(config: RawPlatformConfig) -> dict:
    return {
        "key": config.key,
        "name": config.name,
        "masterTable": config.master_table,
        "sheetName": config.sheet_name,
        "isSupported": config.is_supported,
        "canValidate": config.is_supported,
        "canImport": config.import_handler is not None,
        "canNormalize": config.normalizer is not None,
    }


def _build_header_validation_result(config: RawPlatformConfig, actual_headers: list[str]) -> dict:
    expected_headers = [header for header, _ in config.header_mapping]
    header_errors = config.header_validator(actual_headers)
    extra_headers = actual_headers[len(expected_headers):]

    return {
        "platform": config.key,
        "platformName": config.name,
        "sheetName": config.sheet_name,
        "isCompatible": len(header_errors) == 0,
        "expectedHeaderCount": len(expected_headers),
        "actualHeaderCount": len(actual_headers),
        "headerErrors": header_errors[:50],
        "extraHeaders": extra_headers,
    }


def _get_pending_summary(config: RawPlatformConfig) -> dict:
    if not config.is_supported or config.model is None or config.id_column is None:
        raise ValueError(f"{config.name} import is not supported yet.")

    model = config.model
    id_column = getattr(model, config.id_column)
    pending_filter = model.IsNormalized == False

    pending_raw_rows = (
        db.session.query(func.count(id_column))
        .filter(pending_filter)
        .scalar()
    ) or 0
    pending_active_raw_rows = (
        db.session.query(func.count(id_column))
        .filter(pending_filter)
        .filter(model.isDeleted == False)
        .scalar()
    ) or 0
    pending_deleted_raw_rows = (
        db.session.query(func.count(id_column))
        .filter(pending_filter)
        .filter(model.isDeleted == True)
        .scalar()
    ) or 0
    pending_orders = (
        db.session.query(func.count(func.distinct(model.OrderId)))
        .filter(pending_filter)
        .filter(model.OrderId.isnot(None))
        .filter(model.OrderId != "")
        .scalar()
    ) or 0
    normalize_error_rows = (
        db.session.query(func.count(id_column))
        .filter(pending_filter)
        .filter(model.NormalizeError.isnot(None))
        .filter(model.NormalizeError != "")
        .scalar()
    ) or 0

    latest_pending_row = (
        db.session.query(model)
        .filter(pending_filter)
        .order_by(id_column.desc())
        .first()
    )
    sample_rows = (
        db.session.query(model)
        .filter(pending_filter)
        .order_by(id_column)
        .limit(20)
        .all()
    )

    return {
        "platform": config.key,
        "platformName": config.name,
        "masterTable": config.master_table,
        "pendingRawRows": pending_raw_rows,
        "pendingActiveRawRows": pending_active_raw_rows,
        "pendingDeletedRawRows": pending_deleted_raw_rows,
        "pendingOrders": pending_orders,
        "normalizeErrorRows": normalize_error_rows,
        "latestImportFileName": getattr(latest_pending_row, "ImportFileName", None) if latest_pending_row else None,
        "sampleRows": [row.as_dict() for row in sample_rows],
    }


@router.get("/GetSupportedPlatforms")
async def GetSupportedPlatforms(req: Request):
    try:
        return {
            "status": "success",
            "message": "",
            "data": [_platform_to_dict(config) for config in PLATFORM_CONFIGS.values()],
        }
    except Exception as err:
        error = str(err)
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})


@router.post("/ValidateRawFile")
async def ValidateRawFile(
    req: Request,
    platform: str = Form(...),
    file: UploadFile = File(...),
):
    try:
        config = _get_platform_config(platform)
        if not config.is_supported:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": f"{config.name} import is not supported yet.", "data": ""},
            )

        file_bytes = await file.read()
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        if config.sheet_name not in workbook.sheetnames:
            return {
                "status": "success",
                "message": f"Worksheet '{config.sheet_name}' not found.",
                "data": {
                    "platform": config.key,
                    "platformName": config.name,
                    "fileName": file.filename,
                    "sheetName": config.sheet_name,
                    "isCompatible": False,
                    "availableSheets": workbook.sheetnames,
                },
            }

        worksheet = workbook[config.sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return {
                "status": "success",
                "message": "Excel file has no header row.",
                "data": {
                    "platform": config.key,
                    "platformName": config.name,
                    "fileName": file.filename,
                    "sheetName": config.sheet_name,
                    "isCompatible": False,
                },
            }

        actual_headers = [_normalize_header(value) for value in header_row if value not in (None, "")]
        validation_result = _build_header_validation_result(config, actual_headers)
        validation_result["fileName"] = file.filename
        validation_result["availableSheets"] = workbook.sheetnames

        message = "Raw file is compatible." if validation_result["isCompatible"] else "Raw file headers are not compatible."
        return {"status": "success", "message": message, "data": validation_result}

    except ValueError as err:
        return JSONResponse(status_code=400, content={"status": "error", "message": str(err), "data": ""})
    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/ImportRawFile")
async def ImportRawFile(
    req: Request,
    platform: str = Form(...),
    file: UploadFile = File(...),
    createdBy: str = Form("system"),
):
    try:
        config = _get_platform_config(platform)
        if not config.is_supported or config.import_handler is None:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": f"{config.name} import is not supported yet.", "data": ""},
            )

        result = await config.import_handler(req=req, file=file, createdBy=createdBy)
        return result

    except ValueError as err:
        return JSONResponse(status_code=400, content={"status": "error", "message": str(err), "data": ""})
    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.get("/GetPendingNormalize")
async def GetPendingNormalize(req: Request, platform: str):
    try:
        config = _get_platform_config(platform)
        summary = _get_pending_summary(config)
        return {"status": "success", "message": "", "data": summary}

    except ValueError as err:
        return JSONResponse(status_code=400, content={"status": "error", "message": str(err), "data": ""})
    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/NormalizeRawData")
async def NormalizeRawData(req: Request, body: tb.DataImportNormalizeDTO):
    try:
        config = _get_platform_config(body.platform)
        if not config.is_supported or config.normalizer is None:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": f"{config.name} normalize is not supported yet.", "data": ""},
            )

        if body.mode != "all_pending":
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Only mode='all_pending' is supported.", "data": ""},
            )

        result = config.normalizer(
            limit=None,
            created_by=body.createdBy,
            mode="skip_existing",
        )
        result["platform"] = config.key
        result["platformName"] = config.name
        result["mode"] = body.mode

        return {"status": "success", "message": f"Normalize {config.name} raw data success.", "data": result}

    except ValueError as err:
        db.session.rollback()
        return JSONResponse(status_code=400, content={"status": "error", "message": str(err), "data": ""})
    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()
