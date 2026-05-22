import math
import sys
from datetime import datetime
from io import BytesIO
from os import path

sys.path.append(path.join(path.dirname(__file__), ".."))

from fastapi import APIRouter, Body, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from models import ShopeeMaster
from models.modelsDTO import ShopeeMaster as tb
from config import database as db
from services.ShopeeNormalizer import normalize_shopee_master


router = APIRouter(
    prefix="/ShopeeMaster",
    tags=["ShopeeMaster"],
    responses={
        404: {
            "message": "Not Found"
        }
    }
)

SHMaster = ShopeeMaster.ShopeeMaster

SHOPEE_SHEET_NAME = "raw_shopee"

SHOPEE_HEADER_MAPPING = [
    ("หมายเลขคำสั่งซื้อ", "OrderId"),
    ("สถานะการสั่งซื้อ", "OrderStatus"),
    ("Hot Listing", "HotListing"),
    ("เหตุผลในการยกเลิกคำสั่งซื้อ", "CancellationReason"),
    ("สถานะการคืนเงินหรือคืนสินค้า", "ReturnRefundStatus"),
    ("ชื่อผู้ใช้ (ผู้ซื้อ)", "BuyerUsername"),
    ("วันที่ทำการสั่งซื้อ", "OrderCreatedAt"),
    ("เวลาการชำระสินค้า", "PaidAt"),
    ("ช่องทางการชำระเงิน", "PaymentMethod"),
    ("ช่องทางการชำระเงิน (รายละเอียด)", "PaymentMethodDetail"),
    ("แผนการผ่อนชำระ", "InstallmentPlan"),
    ("ค่าธรรมเนียม (%)", "FeeRate"),
    ("ตัวเลือกการจัดส่ง", "ShippingOption"),
    ("วิธีการจัดส่ง", "ShippingMethod"),
    ("*หมายเลขติดตามพัสดุ", "TrackingNumber"),
    ("วันที่คาดว่าจะทำการจัดส่งสินค้า", "EstimatedShipBy"),
    ("เวลาส่งสินค้า", "ShippedAt"),
    ("เลขอ้างอิง Parent SKU", "ParentSku"),
    ("ชื่อสินค้า", "ProductName"),
    ("เลขอ้างอิง SKU (SKU Reference No.)", "SkuReference"),
    ("ชื่อตัวเลือก", "VariationName"),
    ("ราคาตั้งต้น", "OriginalPrice"),
    ("ราคาขาย", "SalePrice"),
    ("จำนวน", "Quantity"),
    ("จำนวนที่ส่งคืน", "ReturnedQuantity"),
    ("ราคาขายสุทธิ", "NetSalePrice"),
    ("ส่วนลดจาก Shopee", "ShopeeDiscount"),
    ("โค้ดส่วนลดชำระโดยผู้ขาย", "SellerVoucherDiscount"),
    ("โค้ด Coins Cashback ชำระโดยผู้ขาย", "SellerCoinsCashback"),
    ("โค้ดส่วนลดชำระโดย Shopee (เช่น โค้ดจากโปรแกรม ร้านโค้ดคุ้ม, โค้ดส่วนลด Shopee, โค้ดส่วนลด Shopee Mall)", "ShopeeVoucherDiscount"),
    ("โค้ดส่วนลด", "DiscountCodes"),
    ("เข้าร่วมแคมเปญ bundle deal หรือไม่", "IsBundleDeal"),
    ("ส่วนลด bundle deal ชำระโดยผู้ขาย", "SellerBundleDiscount"),
    ("ส่วนลด bundle deal ชำระโดย Shopee", "ShopeeBundleDiscount"),
    ("ส่วนลดจากการใช้เหรียญ", "CoinDiscount"),
    ("โปรโมชั่นช่องทางชำระเงินทั้งหมด", "PaymentChannelPromotionDiscount"),
    ("ส่วนลดเครื่องเก่าแลกใหม่", "TradeInDiscount"),
    ("โบนัสส่วนลดเครื่องเก่าแลกใหม่", "TradeInBonusDiscount"),
    ("ค่าคอมมิชชั่น", "CommissionFee"),
    ("Transaction Fee", "TransactionFee"),
    ("ราคาสินค้าที่ชำระโดยผู้ซื้อ (THB)", "BuyerPaidProductAmountThb"),
    ("ค่าจัดส่งที่ชำระโดยผู้ซื้อ", "BuyerPaidShippingFee"),
    ("ค่าจัดส่งที่ Shopee ออกให้โดยประมาณ", "EstimatedShopeeShippingSubsidy"),
    ("ค่าจัดส่งสินค้าคืน", "ReturnShippingFee"),
    ("ค่าบริการ", "ServiceFee"),
    ("จำนวนเงินทั้งหมด", "TotalAmount"),
    ("ค่าจัดส่งโดยประมาณ", "EstimatedShippingFee"),
    ("โบนัสส่วนลดเครื่องเก่าแลกใหม่จากผู้ขาย", "SellerTradeInBonusDiscount"),
    ("ชื่อผู้รับ", "RecipientName"),
    ("หมายเลขโทรศัพท์", "RecipientPhone"),
    ("หมายเหตุจากผู้ซื้อ", "BuyerNote"),
    ("ที่อยู่ในการจัดส่ง", "ShippingAddress"),
    ("ประเทศ", "ShippingCountry"),
    ("จังหวัด", "ShippingProvince"),
    ("เขต/อำเภอ", "ShippingDistrict"),
    ("รหัสไปรษณีย์", "ShippingPostalCode"),
    ("ประเภทคำสั่งซื้อ", "OrderType"),
    ("เวลาที่ทำการสั่งซื้อสำเร็จ", "CompletedAt"),
    ("บันทึก", "SellerNote"),
    ("ผู้ซื้อร้องขอใบกำกับภาษี", "BuyerRequestedTaxInvoice"),
    ("ประเภทใบกำกับภาษี", "TaxInvoiceType"),
    ("ชื่อ", "TaxInvoiceName"),
    ("ประเภทสาขา", "TaxBranchType"),
    ("ชื่อสาขา", "TaxBranchName"),
    ("รหัสประจำสาขา", "TaxBranchCode"),
    ("ที่อยู่สำหรับออกใบกำกับภาษีแบบเต็มรูป", "TaxFullAddress"),
    ("รายละเอียดที่อยู่", "TaxAddressDetail"),
    ("แขวง/ตำบล", "TaxSubdistrict"),
    ("เขต/อำเภอ", "TaxDistrict"),
    ("จังหวัด", "TaxProvince"),
    ("รหัสไปรษณีย์", "TaxPostalCode"),
    ("หมายเลขประจำตัวผู้เสียภาษี", "TaxId"),
    ("หมายเลขโทรศัพท์สำหรับออกใบกำกับภาษี", "TaxPhone"),
    ("อีเมลสำหรับรับใบกำกับภาษี", "TaxEmail"),
]

def _model_columns():
    return {column.name for column in SHMaster.__table__.columns}

def _writable_columns():
    blocked_columns = {
        "ShopeeMasterId",
        "isDeleted",
        "createdOn",
        "modifiedOn",
    }
    return _model_columns() - blocked_columns

def _payload_to_dict(payload) -> dict:
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=True)
    if hasattr(payload, "dict"):
        return payload.dict(exclude_unset=True)
    return dict(payload)

def _build_model_payload(payload, *, is_create: bool) -> dict:
    payload = _payload_to_dict(payload)
    writable_columns = _writable_columns()
    model_payload = {
        key: value
        for key, value in payload.items()
        if key in writable_columns
    }

    if is_create:
        model_payload["createdBy"] = payload.get("createdBy") or "system"
        model_payload["createdOn"] = datetime.now()
        model_payload["isDeleted"] = False
    else:
        model_payload["modifiedBy"] = payload.get("modifiedBy") or payload.get("createdBy") or "system"
        model_payload["modifiedOn"] = datetime.now()

    return model_payload

def _normalize_header(value):
    return "" if value is None else str(value).strip()

def _cell_to_text(value):
    if value is None:
        return None
    return str(value)

def _validate_shopee_headers(headers: list[str]):
    expected_headers = [header for header, _ in SHOPEE_HEADER_MAPPING]
    actual_known_headers = headers[:len(expected_headers)]

    if len(headers) < len(expected_headers):
        return [
            {
                "column": index + 1,
                "expected": expected_headers[index],
                "actual": actual_known_headers[index] if index < len(actual_known_headers) else None,
            }
            for index in range(len(expected_headers))
            if index >= len(actual_known_headers) or expected_headers[index] != actual_known_headers[index]
        ]

    return [
        {
            "column": index + 1,
            "expected": expected_header,
            "actual": actual_header,
        }
        for index, (expected_header, actual_header) in enumerate(zip(expected_headers, actual_known_headers))
        if expected_header != actual_header
    ]

@router.get("/GetShopeeMaster")
async def GetShopeeMaster(req: Request, page: int, perpage: int = 1000):
    try:
        ShopeeMasterItem = (
            db.session
            .query(SHMaster)
            .filter(SHMaster.isDeleted == False)
            .order_by(SHMaster.ShopeeMasterId)
            .all()
        )

        PaginatedRecord = []
        Page = page
        Perpage = perpage
        TotalRecord = ShopeeMasterItem.__len__()
        AmountPage = 1 if TotalRecord <= Perpage else math.ceil(TotalRecord / Perpage)

        if Page > AmountPage:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Invaild page pagination.", "data": ""},
            )

        TargetIndex = (Page - 1) * Perpage
        FinishElement = TargetIndex + Perpage
        if Page < AmountPage:
            PaginatedRecord = ShopeeMasterItem[TargetIndex:FinishElement]
        else:
            PaginatedRecord = ShopeeMasterItem[TargetIndex:]
        PaginatedRecord = [item.as_dict() for item in PaginatedRecord]

        PaginateReturn = {
            "totalRecords": TotalRecord,
            "pageCount": AmountPage,
            "pageNo": Page,
            "pageSize": Perpage,
        }
        return {
            "status": "success",
            "message": "",
            "data": {"data": PaginatedRecord, "pagination": PaginateReturn},
        }

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.get("/GetShopeeMasterById/{shopeeMasterId}")
async def GetShopeeMasterById(req: Request, shopeeMasterId: int):
    try:
        ShopeeMasterItem = (
            db.session
            .query(SHMaster)
            .filter(SHMaster.ShopeeMasterId == shopeeMasterId)
            .filter(SHMaster.isDeleted == False)
            .first()
        )

        if ShopeeMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "ShopeeMaster not found.", "data": ""},
            )

        return {"status": "success", "message": "", "data": ShopeeMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/ImportShopeeMaster")
async def ImportShopeeMaster(
    req: Request,
    file: UploadFile = File(...),
    createdBy: str = Form("system"),
):
    try:
        file_bytes = await file.read()
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        if SHOPEE_SHEET_NAME not in workbook.sheetnames:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": f"Worksheet '{SHOPEE_SHEET_NAME}' not found.",
                    "data": {"availableSheets": workbook.sheetnames},
                },
            )

        worksheet = workbook[SHOPEE_SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Excel file has no header row.", "data": ""},
            )

        actual_headers = [_normalize_header(value) for value in header_row if value not in (None, "")]
        header_errors = _validate_shopee_headers(actual_headers)
        if header_errors:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": "Shopee Excel headers are not compatible with ShopeeMaster mapping.",
                    "data": {
                        "expectedHeaderCount": len(SHOPEE_HEADER_MAPPING),
                        "actualHeaderCount": len(actual_headers),
                        "headerErrors": header_errors[:30],
                    },
                },
            )

        extra_headers = actual_headers[len(SHOPEE_HEADER_MAPPING):]
        column_names = [column_name for _, column_name in SHOPEE_HEADER_MAPPING]
        created_on = datetime.now()
        records = []
        skipped_empty_rows = 0

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            known_values = row[:len(column_names)]
            if not any(value not in (None, "") for value in known_values):
                skipped_empty_rows += 1
                continue

            model_payload = {
                column_name: _cell_to_text(value)
                for column_name, value in zip(column_names, known_values)
            }
            model_payload["ImportFileName"] = file.filename
            model_payload["IsNormalized"] = False
            model_payload["NormalizedOn"] = None
            model_payload["NormalizeError"] = None
            model_payload["isDeleted"] = False
            model_payload["createdBy"] = createdBy
            model_payload["createdOn"] = created_on
            records.append(SHMaster(**model_payload))

        if records:
            db.session.add_all(records)
            db.session.commit()

        return {
            "status": "success",
            "message": "Import ShopeeMaster success.",
            "data": {
                "fileName": file.filename,
                "sheetName": SHOPEE_SHEET_NAME,
                "importedRows": len(records),
                "skippedEmptyRows": skipped_empty_rows,
                "extraHeaders": extra_headers,
            },
        }

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/NormalizeShopeeMaster")
async def NormalizeShopeeMaster(req: Request, body: tb.ShopeeMasterNormalizeDTO):
    try:
        result = normalize_shopee_master(
            limit=body.limit,
            created_by=body.createdBy,
            mode=body.mode,
        )
        return {"status": "success", "message": "Normalize ShopeeMaster success.", "data": result}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.post("/CreateShopeeMaster")
async def CreateShopeeMaster(req: Request, body: tb.ShopeeMasterCreateDTO):
    try:
        model_payload = _build_model_payload(body, is_create=True)
        ShopeeMasterItem = SHMaster(**model_payload)

        db.session.add(ShopeeMasterItem)
        db.session.commit()
        db.session.refresh(ShopeeMasterItem)

        return {"status": "success", "message": "Create ShopeeMaster success.", "data": ShopeeMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.put("/UpdateShopeeMaster/{shopeeMasterId}")
async def UpdateShopeeMaster(req: Request, shopeeMasterId: int, body: tb.ShopeeMasterUpdateDTO):
    try:
        ShopeeMasterItem = (
            db.session
            .query(SHMaster)
            .filter(SHMaster.ShopeeMasterId == shopeeMasterId)
            .filter(SHMaster.isDeleted == False)
            .first()
        )

        if ShopeeMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "ShopeeMaster not found.", "data": ""},
            )

        model_payload = _build_model_payload(body, is_create=False)
        for key, value in model_payload.items():
            setattr(ShopeeMasterItem, key, value)

        db.session.commit()
        db.session.refresh(ShopeeMasterItem)

        return {"status": "success", "message": "Update ShopeeMaster success.", "data": ShopeeMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()


@router.delete("/DeleteShopeeMaster/{shopeeMasterId}")
async def DeleteShopeeMaster(req: Request, shopeeMasterId: int, payload: dict = Body(default={})):
    try:
        ShopeeMasterItem = (
            db.session
            .query(SHMaster)
            .filter(SHMaster.ShopeeMasterId == shopeeMasterId)
            .filter(SHMaster.isDeleted == False)
            .first()
        )

        if ShopeeMasterItem is None:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "ShopeeMaster not found.", "data": ""},
            )

        ShopeeMasterItem.isDeleted = True
        ShopeeMasterItem.modifiedBy = payload.get("modifiedBy") or payload.get("createdBy") or "system"
        ShopeeMasterItem.modifiedOn = datetime.now()

        db.session.commit()
        db.session.refresh(ShopeeMasterItem)

        return {"status": "success", "message": "Delete ShopeeMaster success.", "data": ShopeeMasterItem.as_dict()}

    except Exception as err:
        error = str(err)
        db.session.rollback()
        return JSONResponse(status_code=500, content={"status": "error", "message": error, "data": ""})
    finally:
        db.session.close()
